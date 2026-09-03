"""Core logic for parsing and merging POS decline reports.

Accepts any number of Excel files (.xls / .xlsx) that contain a POS decline
transaction table (a sheet whose header row contains "ACQUIRER") and merges
them into a single workbook laid out like POS_Transaction_Decline_Report,
with blank/spacer columns removed. Everything runs in memory - nothing is
written to disk.
"""

from __future__ import annotations

import html
import io
import re
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Report modes: POS decline, ATM and POS transaction reports each have
# their own canonical column set (defined by the *sample output report*),
# header aliases and output layout. Columns present in an uploaded file
# but NOT in the sample output format are removed from the output.
# ---------------------------------------------------------------------------
# POS decline mode -> POS_Transaction_Decline_Report format
POS_DECLINE_CANONICAL_COLUMNS = (
    "ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
    "AMOUNT", "RESP_CODE", "Reversal", "FE UTRNNO", "REFNUM", "MERCHANT",
)

POS_DECLINE_HEADER_ALIASES = {
    "ACQUIRER": "ACQUIRER",
    "ISSUER": "ISSUER",
    "PAN": "PAN",
    "TRANS_DATE": "TRAN_DATE",
    "TRAN_DATE": "TRAN_DATE",
    "TRANS DATE": "TRAN_DATE",
    "TRAN DATE": "TRAN_DATE",
    "TRANSACTION DATE": "TRAN_DATE",
    "TIME": "TIME",
    "TRANS_TYPE": "TRANS_TYPE",
    "AMOUNT": "AMOUNT",
    "RESP": "RESP_CODE",
    "RESP_CODE": "RESP_CODE",
    "RESP CODE": "RESP_CODE",
    "REVERSAL": "Reversal",
    "FE_UTRNNO": "FE UTRNNO",
    "FE UTRNNO": "FE UTRNNO",
    "UTRNNO": "FE UTRNNO",
    "UTRN NO": "FE UTRNNO",
    "REFNUM": "REFNUM",
    "REF NUM": "REFNUM",
    "ADDRESS_NAME": "MERCHANT",
    "ADDRESS NAME": "MERCHANT",
    "MERCHANT": "MERCHANT",
    "MERCHANT NAME": "MERCHANT",
}

# POS mode -> Daily_Tranaction_Report_SmartVista_POS format
# (sample file: Daily_Tranaction_Report_SmartVista_POS_15_Aug_26_to_15_Aug_26.xlsx)
POS_CANONICAL_COLUMNS = (
    "ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
    "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN", "TERMINAL_ID",
    "ADDRESS",
)

POS_HEADER_ALIASES = {
    "ACQUIRER": "ACQUIRER",
    "ISSUER": "ISSUER",
    "CARD_NUMBER": "CARD_NUMBER",
    "CARD NO": "CARD_NUMBER",
    "CARD NO.": "CARD_NUMBER",
    "PAN": "CARD_NUMBER",
    "TRANS_DATE": "TRANS_DATE",
    "TRAN_DATE": "TRANS_DATE",
    "TRANS DATE": "TRANS_DATE",
    "TRAN DATE": "TRANS_DATE",
    "TRANSACTION DATE": "TRANS_DATE",
    "TRANS_TIME": "TRANS_TIME",
    "TIME": "TRANS_TIME",
    "TRANS TIME": "TRANS_TIME",
    "TRANS_TYPE": "TRANS_TYPE",
    "AMOUNT": "AMOUNT",
    "CURRENCY": "CURRENCY",
    "CURRENCY CODE": "CURRENCY",
    "CUR": "CURRENCY",
    "RESP": "RESP",
    "RESP_CODE": "RESP",
    "RESP CODE": "RESP",
    "RRN": "RRN",
    "RETRIEVAL REFERENCE NUMBER": "RRN",
    # UTRNNO is present in some POS exports but NOT in the POS output
    # sample - it is recognized so it can be reported as removed.
    "UTRNNO": "UTRNNO",
    "UTRN NO": "UTRNNO",
    "FE UTRNNO": "UTRNNO",
    "FE_UTRNNO": "UTRNNO",
    "TERMINAL_ID": "TERMINAL_ID",
    "TERMINAL ID": "TERMINAL_ID",
    "ATM ID": "TERMINAL_ID",
    "ADDRESS": "ADDRESS",
    "ADDRESS_NAME": "ADDRESS",
    "ADDRESS NAME": "ADDRESS",
    "MERCHANT": "ADDRESS",
    "MERCHANT NAME": "ADDRESS",
}

# POS success mode -> POS_Transaction_SVFE_Report format
POS_SUCCESS_CANONICAL_COLUMNS = (
    "ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
    "AMOUNT", "RESP_CODE", "REFNUM", "UTRNNO", "MERCHANT",
)

POS_SUCCESS_HEADER_ALIASES = {
    "ACQUIRER": "ACQUIRER",
    "ISSUER": "ISSUER",
    "PAN": "PAN",
    "CARD_NUMBER": "PAN",
    "CARD NO": "PAN",
    "CARD NO.": "PAN",
    "TRANS_DATE": "TRAN_DATE",
    "TRAN_DATE": "TRAN_DATE",
    "TRANS DATE": "TRAN_DATE",
    "TRAN DATE": "TRAN_DATE",
    "TRANSACTION DATE": "TRAN_DATE",
    "TIME": "TIME",
    "TRANS_TIME": "TIME",
    "TRANS TYPE": "TRANS_TYPE",
    "TRANS_TYPE": "TRANS_TYPE",
    "AMOUNT": "AMOUNT",
    "RESP": "RESP_CODE",
    "RESP_CODE": "RESP_CODE",
    "RESP CODE": "RESP_CODE",
    "REVERSAL": "REVERSAL",
    "REFNUM": "REFNUM",
    "REF NUM": "REFNUM",
    "FE_UTRNNO": "UTRNNO",
    "FE UTRNNO": "UTRNNO",
    "UTRNNO": "UTRNNO",
    "UTRN NO": "UTRNNO",
    "ADDRESS_NAME": "MERCHANT",
    "ADDRESS NAME": "MERCHANT",
    "MERCHANT": "MERCHANT",
    "MERCHANT NAME": "MERCHANT",
}

# ATM mode -> Daily_Tranaction_Report_SmartVista_ATM format
ATM_CANONICAL_COLUMNS = (
    "ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
    "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN", "UTRNNO",
    "TERMINAL_ID", "ADDRESS_NAME",
)

ATM_HEADER_ALIASES = {
    "ACQUIRER": "ACQUIRER",
    "ISSUER": "ISSUER",
    "CARD_NUMBER": "CARD_NUMBER",
    "CARD NO": "CARD_NUMBER",
    "CARD NO.": "CARD_NUMBER",
    "PAN": "CARD_NUMBER",
    "TRANS_DATE": "TRANS_DATE",
    "TRAN_DATE": "TRANS_DATE",
    "TRANS DATE": "TRANS_DATE",
    "TRAN DATE": "TRANS_DATE",
    "TRANSACTION DATE": "TRANS_DATE",
    "TRANS_TIME": "TRANS_TIME",
    "TIME": "TRANS_TIME",
    "TRANS TIME": "TRANS_TIME",
    "TRANS_TYPE": "TRANS_TYPE",
    "AMOUNT": "AMOUNT",
    "CURRENCY": "CURRENCY",
    "CURRENCY CODE": "CURRENCY",
    "CUR": "CURRENCY",
    "RESP": "RESP",
    "RESP_CODE": "RESP",
    "RESP CODE": "RESP",
    "RRN": "RRN",
    "RETRIEVAL REFERENCE NUMBER": "RRN",
    "UTRNNO": "UTRNNO",
    "UTRN NO": "UTRNNO",
    "FE UTRNNO": "UTRNNO",
    "FE_UTRNNO": "UTRNNO",
    "TERMINAL_ID": "TERMINAL_ID",
    "TERMINAL ID": "TERMINAL_ID",
    "ATM ID": "TERMINAL_ID",
    "ADDRESS_NAME": "ADDRESS_NAME",
    "ADDRESS NAME": "ADDRESS_NAME",
    "MERCHANT": "ADDRESS_NAME",
    "MERCHANT NAME": "ADDRESS_NAME",
}

# QR transfer/export mode -> EXPORT_TABLE format ("July - December 2025
# Source"). The columns come straight from the interbank transfer export:
# DESTINATION_BANK, SOURCE_BANK, TRX_DATE, DBTR_ACCT (debtor account),
# CDTR_ACCT (creditor account), AMOUNT, TX_ID (transaction id) and STATUS.
QR_CANONICAL_COLUMNS = (
    "DESTINATION_BANK", "SOURCE_BANK", "TRX_DATE", "DBTR_ACCT",
    "CDTR_ACCT", "AMOUNT", "TX_ID", "STATUS",
)

QR_HEADER_ALIASES = {
    "DESTINATION_BANK": "DESTINATION_BANK",
    "DESTINATION BANK": "DESTINATION_BANK",
    "DEST BANK": "DESTINATION_BANK",
    "RECEIVER BANK": "DESTINATION_BANK",
    "BENEFICIARY BANK": "DESTINATION_BANK",
    "SOURCE_BANK": "SOURCE_BANK",
    "SOURCE BANK": "SOURCE_BANK",
    "SENDING BANK": "SOURCE_BANK",
    "ORIGINATOR BANK": "SOURCE_BANK",
    "TRX_DATE": "TRX_DATE",
    "TRX DATE": "TRX_DATE",
    "TRANSACTION DATE": "TRX_DATE",
    "TRANS DATE": "TRX_DATE",
    "TRAN DATE": "TRX_DATE",
    "TRANS_DATE": "TRX_DATE",
    "TRAN_DATE": "TRX_DATE",
    "DATE": "TRX_DATE",
    "DBTR_ACCT": "DBTR_ACCT",
    "DBTR ACCT": "DBTR_ACCT",
    "DEBTOR ACCOUNT": "DBTR_ACCT",
    "DEBTOR ACCT": "DBTR_ACCT",
    "DEBIT ACCOUNT": "DBTR_ACCT",
    "DEBIT ACCT": "DBTR_ACCT",
    "SENDER ACCOUNT": "DBTR_ACCT",
    "SENDER ACCT": "DBTR_ACCT",
    "FROM ACCOUNT": "DBTR_ACCT",
    "CDTR_ACCT": "CDTR_ACCT",
    "CDTR ACCT": "CDTR_ACCT",
    "CREDITOR ACCOUNT": "CDTR_ACCT",
    "CREDITOR ACCT": "CDTR_ACCT",
    "CREDIT ACCOUNT": "CDTR_ACCT",
    "CREDIT ACCT": "CDTR_ACCT",
    "RECEIVER ACCOUNT": "CDTR_ACCT",
    "RECEIVER ACCT": "CDTR_ACCT",
    "TO ACCOUNT": "CDTR_ACCT",
    "AMOUNT": "AMOUNT",
    "TRANSACTION AMOUNT": "AMOUNT",
    "TX AMOUNT": "AMOUNT",
    "AMOUNT (ETB)": "AMOUNT",
    "AMOUNT - ETB": "AMOUNT",
    "TX_ID": "TX_ID",
    "TX ID": "TX_ID",
    "TRANSACTION ID": "TX_ID",
    "TRANSACTION REFERENCE": "TX_ID",
    "TRANSACTION REF": "TX_ID",
    "TX REF": "TX_ID",
    "REFERENCE": "TX_ID",
    "TRANSACTION NO": "TX_ID",
    "STATUS": "STATUS",
    "TRANSACTION STATUS": "STATUS",
    "TX STATUS": "STATUS",
    "RESPONSE": "STATUS",
}

# Backward-compatible names (POS decline mode defaults)
CANONICAL_COLUMNS = list(POS_DECLINE_CANONICAL_COLUMNS)
HEADER_ALIASES = dict(POS_DECLINE_HEADER_ALIASES)

REPORT_TITLE = "POS TRANSACTION  DECLINE REPORT"
SHEET_NAME = "POS_Transaction_Report"

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _pos_file_date(yyyymmdd: str) -> str:
    """'20260813' -> '20260813' (plain YYYYMMDD)."""
    return yyyymmdd


def _atm_file_date(yyyymmdd: str) -> str:
    """'20260815' -> '15_Aug_26' (as in the SmartVista ATM report name)."""
    if len(yyyymmdd) == 8 and yyyymmdd.isdigit():
        y, m, d = int(yyyymmdd[:4]), int(yyyymmdd[4:6]), int(yyyymmdd[6:8])
        if 1 <= m <= 12:
            return f"{d:02d}_{_MONTHS[m - 1]}_{str(y)[2:]}"
    return yyyymmdd


@dataclass(frozen=True)
class ReportMode:
    key: str
    label: str
    canonical_columns: tuple
    header_aliases: dict
    sheet_name: str
    report_title: str | None
    output_prefix: str
    sample_label: str          # name of the sample output report (for warnings)
    title_rows: int            # title rows before the header row (0 = none)
    column_widths: dict
    numeric_fmt_cols: tuple
    resp_column: str
    date_column: str
    time_column: str
    file_date: Any             # callable: YYYYMMDD str -> filename date str
    always_show_range: bool


POS_SUCCESS_MODE = ReportMode(
    key="pos_success",
    label="POS Success",
    canonical_columns=POS_SUCCESS_CANONICAL_COLUMNS,
    header_aliases=POS_SUCCESS_HEADER_ALIASES,
    sheet_name="POS_Transaction_Report",
    report_title="POS TRANSACTION REPORT",
    output_prefix="POS_Transaction_SVFE_Report",
    sample_label="POS_Transaction_SVFE_Report",
    title_rows=3,
    column_widths={
        "A": 16, "B": 14, "C": 22, "D": 12, "E": 10, "F": 14,
        "G": 12, "H": 11, "I": 16, "J": 20, "K": 42,
    },
    numeric_fmt_cols=("TRAN_DATE", "RESP_CODE", "REFNUM", "UTRNNO"),
    resp_column="RESP_CODE",
    date_column="TRAN_DATE",
    time_column="TIME",
    file_date=_pos_file_date,
    always_show_range=False,
)

POS_DECLINE_MODE = ReportMode(
    key="pos_decline",
    label="POS Decline",
    canonical_columns=POS_DECLINE_CANONICAL_COLUMNS,
    header_aliases=POS_DECLINE_HEADER_ALIASES,
    sheet_name="POS_Transaction_Report",
    report_title="POS TRANSACTION  DECLINE REPORT",
    output_prefix="POS_Transaction_Decline_Report",
    sample_label="POS_Transaction_Decline_Report",
    title_rows=3,
    column_widths={
        "A": 16, "B": 14, "C": 22, "D": 12, "E": 10, "F": 14,
        "G": 12, "H": 11, "I": 9, "J": 20, "K": 16, "L": 42,
    },
    numeric_fmt_cols=("TRAN_DATE", "RESP_CODE", "Reversal", "FE UTRNNO", "REFNUM"),
    resp_column="RESP_CODE",
    date_column="TRAN_DATE",
    time_column="TIME",
    file_date=_pos_file_date,
    always_show_range=False,
)

POS_MODE = ReportMode(
    key="pos",
    label="POS",
    canonical_columns=POS_CANONICAL_COLUMNS,
    header_aliases=POS_HEADER_ALIASES,
    sheet_name="Report",
    report_title=None,
    output_prefix="Daily_Tranaction_Report_SmartVista_POS",
    sample_label="Daily_Tranaction_Report_SmartVista_POS",
    title_rows=0,
    column_widths={
        "A": 18, "B": 16, "C": 22, "D": 12, "E": 12, "F": 22,
        "G": 12, "H": 10, "I": 9, "J": 18, "K": 13, "L": 42,
    },
    numeric_fmt_cols=("TRANS_DATE", "RESP", "RRN"),
    resp_column="RESP",
    date_column="TRANS_DATE",
    time_column="TRANS_TIME",
    file_date=_atm_file_date,
    always_show_range=True,
)

ATM_MODE = ReportMode(
    key="atm",
    label="ATM",
    canonical_columns=ATM_CANONICAL_COLUMNS,
    header_aliases=ATM_HEADER_ALIASES,
    sheet_name="Report",
    report_title=None,
    output_prefix="Daily_Tranaction_Report_SmartVista_ATM",
    sample_label="Daily_Tranaction_Report_SmartVista_ATM",
    title_rows=0,
    column_widths={
        "A": 18, "B": 16, "C": 22, "D": 12, "E": 12, "F": 22,
        "G": 12, "H": 10, "I": 9, "J": 18, "K": 22, "L": 13, "M": 42,
    },
    numeric_fmt_cols=("TRANS_DATE", "RESP", "RRN", "UTRNNO"),
    resp_column="RESP",
    date_column="TRANS_DATE",
    time_column="TRANS_TIME",
    file_date=_atm_file_date,
    always_show_range=True,
)

# QR transfer-export mode -> "July - December 2025 Source" EXPORT_TABLE
# format. It is a plain tabular export (header in row 1, no title block),
# so it uses the streaming (SmartVista-style) writer. The whole transaction
# date+time lives in TRX_DATE, so date-then-time ordering collapses to a
# date sort and STATUS ("PROCESSED"/"DECLINED") drives the distribution.
QR_MODE = ReportMode(
    key="qr",
    label="QR",
    canonical_columns=QR_CANONICAL_COLUMNS,
    header_aliases=QR_HEADER_ALIASES,
    sheet_name="Report",
    report_title=None,
    output_prefix="QR_Export",
    sample_label="QR_Export",
    title_rows=0,
    column_widths={
        "A": 22, "B": 24, "C": 16, "D": 20, "E": 20,
        "F": 12, "G": 24, "H": 12,
    },
    numeric_fmt_cols=(),
    resp_column="STATUS",
    date_column="TRX_DATE",
    time_column="TRX_DATE",
    file_date=_atm_file_date,
    always_show_range=True,
)

MODES = {
    "pos_decline": POS_DECLINE_MODE,
    "pos_success": POS_SUCCESS_MODE,
    "pos": POS_MODE,
    "atm": ATM_MODE,
    "qr": QR_MODE,
}

# Canonical column names that are EXCLUDED from the duplicate-row fingerprint
# across ALL report modes (POS Decline, POS Success, POS Daily, ATM, QR).
#
# ACQUIRER, ISSUER, TRANS_TYPE and CURRENCY do not uniquely identify a
# transaction - the same card swipe / ATM withdrawal can appear in exports
# from different acquirers or with a slightly different TRANS_TYPE label.
# Excluding them means two rows are treated as duplicates when every other
# column (card/account number, date, time, amount, response code, reference
# numbers, terminal ID, merchant/address) matches.
#
# Applied uniformly: if a mode does not have one of these columns (e.g. QR
# has no CURRENCY or TRANS_TYPE) the missing column is simply not present in
# the record, so the exclusion has no effect on those modes.
DUPLICATE_IGNORE_COLUMNS: frozenset[str] = frozenset({
    "ACQUIRER",
    "ISSUER",
    "TRANS_TYPE",
    "CURRENCY",
})


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------
# NOTE: values are NEVER altered. The helpers below only classify emptiness,
# build lookup keys, or produce internal sort/count keys - the raw value as
# read from the file (including literal "null" strings, raw TIME formats and
# original number types) is what ends up in the merged output.
def _is_empty(value: Any) -> bool:
    if value is None or value is pd.NA:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    # pandas returns '' for empty cells (keep_default_na=False); whitespace-
    # only cells are treated as empty too, like Excel does.
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _cell_value(value: Any) -> Any:
    """Raw cell value; empty cells become '' (never altered otherwise)."""
    return "" if _is_empty(value) else value


def _count_key(value: Any) -> str:
    """Grouping key for the response-code chart only - never stored.
    901, 901.0 and '901' all count as the same response code."""
    if _is_empty(value):
        return ""
    try:
        f = float(value)
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return str(value)


def _normalize_date(value: Any) -> str:
    """Return the transaction date as a 'YYYYMMDD' string."""
    if _is_empty(value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        v = int(value)
        return str(v) if 19000101 <= v <= 99991231 else str(v)
    s = str(value).strip()
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$", s)
    if m:
        return f"{m.group(1)}{int(m.group(2)):02d}{int(m.group(3)):02d}"
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})$", s)
    if m:
        return f"{m.group(3)}{int(m.group(1)):02d}{int(m.group(2)):02d}"
    digits = re.sub(r"\D", "", s)
    return digits if len(digits) == 8 else s


def _normalize_time(value: Any) -> str:
    """Normalize a TIME cell to 'HH:MM:SS'.

    Accepts raw HHMMSS of length 1..6 (number or string) or an already
    formatted 'HH:MM' / 'HH:MM:SS' string. Mirrors the CASE logic used by
    the source SQL that generates the reports.
    """
    if _is_empty(value):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value = str(int(value))
    s = str(value).strip()
    if not s:
        return ""
    if ":" in s:
        parts = s.split(":")
        if len(parts) == 2:
            return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}:00"
        # zero-pad every segment so mixed formats ("0:0:20" vs "00:24:22")
        # sort chronologically; this key is used for ordering only
        return ":".join(p.zfill(2) for p in parts[:3])
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    digits = digits.zfill(6)
    return f"{digits[0:2]}:{digits[2:4]}:{digits[4:6]}"


_MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _textual_date_to_ymd(s: str) -> tuple[int, int, int] | None:
    """Parse a date written with a spelled-out month into (year, month, day).

    Handles formats like '15-Aug-26', 'Aug 15, 2026', '13 Aug 2026',
    '15 AUGUST 26', '15-August-2026', 'Aug 15 26', etc. The year may be
    two digits (assumed to belong to the 2000s) or four digits. Returns
    None if no spelled month is found or the parts do not make a date.
    """
    lower = s.lower()
    month = None
    month_name = None
    for name, num in _MONTH_NAMES.items():
        if name in lower:
            month = num
            month_name = name
            break
    if not month:
        return None
    rest = re.sub(r"[^a-z0-9]+", " ", lower)
    tokens = [t for t in rest.split() if t]
    # remove the (abbreviated or full) month token
    tokens = [
        t for t in tokens
        if not (t.startswith(month_name) or month_name.startswith(t))
    ]
    nums = [int(t) for t in tokens if t.isdigit()]
    if not nums:
        return None
    day = None
    year = None
    if len(nums) == 1:
        n = nums[0]
        if n > 31:
            year = n
        else:
            day = n
    else:
        # two numbers: one is the day (<=31), the other the year
        n1, n2 = nums[0], nums[1]
        if n1 > 31:
            year, day = n1, n2
        elif n2 > 31 or n2 > 99:
            day, year = n1, n2
        else:
            # both small: typical "15 Aug 26" -> first is day, second is year
            day, year = n1, n2
    if not day:
        return None
    if year is None:
        year = 2000
    if year < 100:
        year += 2000
    if 1 <= month <= 12 and 1 <= day <= 31:
        return (year, month, day)
    return None


def smart_sort_key(value: Any) -> tuple:
    """Build a comparable sort key for an arbitrary cell value.

    Handles mixed representations so that a column containing dates written
    with spelled months (e.g. '15-Aug-26', 'Aug 13, 2026'), numeric dates
    (e.g. 20260813), times, numbers and plain text all sort correctly within
    the same column. The returned tuple puts every value in the same
    'namespace' so the ordering is stable and deterministic.
    """
    if _is_empty(value):
        return (1, 0, "", "", 0)  # empty values always sort first

    # datetime / date objects
    if isinstance(value, datetime):
        return (2, 0, "", value.isoformat(), 0)
    if isinstance(value, date):
        return (2, 0, "", value.isoformat(), 0)

    s = str(value).strip()

    # textual date with a spelled month (highest priority after objects)
    ymd = _textual_date_to_ymd(s)
    if ymd:
        return (2, 0, "", f"{ymd[0]:04d}-{ymd[1]:02d}-{ymd[2]:02d}", 0)

    # time-like values (contain ':' or look like raw HHMMSS)
    if ":" in s:
        t = _normalize_time(value)
        if t:
            return (3, 0, "", t, 0)

    # 8-digit YYYYMMDD dates (numeric - e.g. 20260813) sort as dates, so they
    # interleave correctly with dates written using spelled months.
    if len(s) == 8 and s.isdigit():
        return (2, 0, "", f"{s[0:4]}-{s[4:6]}-{s[6:8]}", 0)

    # pure numbers
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (4, float(value), "", "", 0)
    num = re.fullmatch(r"[+-]?\d+(?:\.\d+)?", s)
    if num:
        return (4, float(s), "", "", 0)

    # plain text
    return (5, 0, "", s.lower(), 0)


def sort_records(records: list[dict], column: str = "", descending: bool = False) -> list[dict]:
    """Return *records* sorted by the given column using smart keys.

    If ``column`` is empty/unknown the list is returned in its original
    order. ``descending`` reverses the order (empty values keep sorting
    first regardless, so blank rows stay at the top).
    """
    if not column:
        return list(records)
    ordered = sorted(records, key=lambda r: smart_sort_key(r.get(column)))
    if descending:
        # keep empty rows pinned to the front, reverse the rest
        empty = [r for r in ordered if _is_empty(r.get(column))]
        filled = [r for r in ordered if not _is_empty(r.get(column))]
        filled.reverse()
        ordered = empty + filled
    return ordered


def _reverse_date_time_sort(records: list[dict], mode: ReportMode) -> None:
    """Reverse an existing date-then-time sorted list in place, keeping rows
    with no date pinned to the front and preserving the original relative
    order among undated rows."""
    chunks: list[list] = []
    current: list = []
    for r in records:
        if _is_empty(r.get(mode.date_column)):
            if current:
                chunks.append(current)
                current = []
            chunks.append([r])  # each undated row stays as its own unit
        else:
            current.append(r)
    if current:
        chunks.append(current)
    new_records: list = []
    for chunk in reversed(chunks):
        if chunk and _is_empty(chunk[0].get(mode.date_column)):
            new_records.extend(chunk)
        else:
            chunk.reverse()
            new_records.extend(chunk)
    records[:] = new_records


def _detect_engine(data: bytes) -> str:
    """Pick the right pandas engine from the file signature."""
    if data[:2] == b"PK":  # xlsx / zip container
        return "openpyxl"
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # OLE2 / legacy xls
        return "xlrd"
    raise ValueError("unsupported file format (expected an .xls or .xlsx file)")


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def _header_markers(mode: ReportMode) -> set[str]:
    """The distinctive header cell(s) that mark a header row for a mode.

    POS-family reports always carry an 'ACQUIRER' header; QR transfer
    exports carry 'DESTINATION_BANK' (and SOURCE_BANK / TRX_DATE). A header
    row is recognized when any of its cells matches one of these tokens.
    """
    if mode is not None and mode.key == "qr":
        return {"DESTINATION_BANK", "SOURCE_BANK", "TRX_DATE"}
    return {"ACQUIRER"}


def _is_header_row(row, mode: ReportMode | None = None) -> bool:
    """A header row is any row containing a distinctive header cell for the
    mode (e.g. 'ACQUIRER' for POS-family, 'DESTINATION_BANK' for QR)."""
    if mode is not None and mode.key == "qr":
        # QR exports may spell headers descriptively ("Destination Bank",
        # "Debit Acct", "Transaction ID", ...), so a row counts as a header
        # when several of its cells map by name to canonical QR columns.
        mapped = 0
        for v in row:
            if _is_empty(v):
                continue
            canon = mode.header_aliases.get(str(v).strip().upper())
            if canon and canon in mode.canonical_columns:
                mapped += 1
        return mapped >= 3
    markers = _header_markers(mode)
    for v in row:
        if _is_empty(v):
            continue
        if str(v).strip().upper() in markers:
            return True
    return False


# Files above this size are read with the fast regex-based xlsx parser
# (several times faster than openpyxl); smaller files use openpyxl, which is
# battle-tested and fast enough at that size. Both parsers produce identical
# raw values - numbers are parsed from their exact decimal strings, so large
# UTRNNO values keep full precision.
_FAST_READ_MIN_BYTES = 8 * 1024 * 1024


def _read_grids(data: bytes, engine: str) -> dict:
    """Read every sheet as a list-of-lists of raw cell values.

    xlsx is read with either a fast regex-based parser (large files) or
    openpyxl in read-only mode (small files). Legacy .xls is read with
    pandas/xlrd and converted to the same raw structure.
    """
    if engine == "openpyxl":
        if len(data) >= _FAST_READ_MIN_BYTES:
            try:
                grids = _read_xlsx_fast(data)
                if grids:
                    return grids
            except Exception:  # noqa: BLE001 - fall back to openpyxl
                pass
        return _read_xlsx_openpyxl(data)

    sheets = pd.read_excel(
        io.BytesIO(data),
        sheet_name=None,
        header=None,
        dtype=object,
        keep_default_na=False,
        engine="xlrd",
    )
    return {
        name: [[None if _is_empty(v) else v for v in row] for row in grid.values.tolist()]
        for name, grid in sheets.items()
    }


def _read_xlsx_openpyxl(data: bytes) -> dict:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return {
            ws.title: [list(row) for row in ws.iter_rows(values_only=True)]
            for ws in wb.worksheets
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Fast xlsx reader (regex / string-level, no element objects).
# The worksheet XML is flat (<row>...<c>...</c>...</row>), so it can be
# parsed with a handful of regular expressions at string level, which is
# several times faster than openpyxl for reports with hundreds of thousands
# of rows. Cell values are converted exactly like openpyxl does (ints stay
# ints, decimals/scientific-notation become floats), so nothing is altered.
# ---------------------------------------------------------------------------
_XML_TAG = r"(?:\w+:)?"  # matches both <row> and <x:row>
_ROW_BLOCK_RE = re.compile(rf"<{_XML_TAG}row\b[^>]*>(.*?)</{_XML_TAG}row>", re.S)
_CELL_BLOCK_RE = re.compile(rf"<{_XML_TAG}c\b([^>]*?)(?:/>|>(.*?)</{_XML_TAG}c>)", re.S)
_SI_BLOCK_RE = re.compile(rf"<{_XML_TAG}si\b[^>]*>(.*?)</{_XML_TAG}si>", re.S)
_SHEET_BLOCK_RE = re.compile(rf"<{_XML_TAG}sheet\b[^>]*>", re.S)
_REL_BLOCK_RE = re.compile(rf"<{_XML_TAG}Relationship\b[^>]*>", re.S)
_ATTR_RE = re.compile(r'([A-Za-z]+(?::[A-Za-z]+)?)="([^"]*)"')
_VAL_RE = re.compile(rf"<{_XML_TAG}v>([^<]*)</{_XML_TAG}v>")
_TEXT_RE = re.compile(rf"<{_XML_TAG}t(?:\s[^>]*)?>(.*?)</{_XML_TAG}t>", re.S)
_IS_RE = re.compile(rf"<{_XML_TAG}is>(.*?)</{_XML_TAG}is>", re.S)
_COL_REF_RE = re.compile(r"([A-Z]+)")


def _xml_attrs(block: str) -> dict:
    return dict(_ATTR_RE.findall(block))


def _xml_unescape(text: str) -> str:
    """Decode XML entities, including numeric character references like
    &#233; (openpyxl escapes non-ASCII characters this way)."""
    return html.unescape(text)


def _col_ref_to_index(ref: str) -> int:
    """'AB12' -> 27 (0-based column index)."""
    n = 0
    for ch in _COL_REF_RE.match(ref).group(1):
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _parse_xml_cell(attrs: str, content: str | None, shared: list) -> Any:
    am = _xml_attrs(attrs)
    t = am.get("t")
    if t == "s":
        vm = _VAL_RE.search(content or "")
        return shared[int(vm.group(1))] if vm else None
    if t == "inlineStr":
        ism = _IS_RE.search(content or "")
        if not ism:
            return None  # empty inline string cell (no <is> child)
        return _xml_unescape("".join(_TEXT_RE.findall(ism.group(1))))
    if t == "str":
        vm = _VAL_RE.search(content or "")
        return _xml_unescape(vm.group(1)) if vm else ""
    if t == "b":
        vm = _VAL_RE.search(content or "")
        return vm.group(1) == "1" if vm else None
    if t == "e":
        vm = _VAL_RE.search(content or "")
        return vm.group(1) if vm else ""
    vm = _VAL_RE.search(content or "")
    if not vm:
        return None
    txt = vm.group(1)
    return float(txt) if ("." in txt or "E" in txt or "e" in txt) else int(txt)


def _read_xlsx_fast(data: bytes) -> dict:
    """Parse an .xlsx file from its raw XML at string level."""
    zf = zipfile.ZipFile(io.BytesIO(data))
    names = set(zf.namelist())

    # shared strings table (strings referenced by index from the cells)
    shared: list[str] = []
    if "xl/sharedStrings.xml" in names:
        sxml = zf.read("xl/sharedStrings.xml").decode("utf-8-sig", "replace")
        shared = [
            _xml_unescape("".join(_TEXT_RE.findall(si)))
            for si in _SI_BLOCK_RE.findall(sxml)
        ]

    wbxml = zf.read("xl/workbook.xml").decode("utf-8-sig", "replace")
    sheets = []
    for m in _SHEET_BLOCK_RE.finditer(wbxml):
        am = _xml_attrs(m.group(0))
        name, rid = am.get("name"), am.get("r:id")
        if name is not None and rid is not None:
            sheets.append((name, rid))

    relxml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8-sig", "replace")
    rels = {
        _xml_attrs(m.group(0)).get("Id"): _xml_attrs(m.group(0)).get("Target")
        for m in _REL_BLOCK_RE.finditer(relxml)
    }

    grids: dict[str, list] = {}
    for name, rid in sheets:
        target = rels.get(rid)
        if not target:
            continue
        # Targets may be absolute ("/xl/worksheets/sheet1.xml") or relative
        # ("worksheets/sheet1.xml") - both resolve to xl/worksheets/sheet1.xml
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        try:
            xml = zf.read(target).decode("utf-8-sig", "replace")
        except KeyError:
            continue
        grid = []
        for row_body in _ROW_BLOCK_RE.findall(xml):
            row: list[Any] = []
            idx = 0
            for attrs, content in _CELL_BLOCK_RE.findall(row_body):
                am = _xml_attrs(attrs)
                ref = am.get("r")
                if ref:
                    idx = _col_ref_to_index(ref)
                while len(row) < idx:
                    row.append(None)
                row.append(_parse_xml_cell(attrs, content, shared))
                idx += 1
            grid.append(row)
        grids[name] = grid
    return grids


@dataclass
class FileReport:
    filename: str
    sheet: str = ""
    header_row: int = 0
    raw_rows: int = 0
    data_rows: int = 0
    columns_kept: list = field(default_factory=list)
    # Canonical column names in the order they appear in the source sheet
    column_order: list = field(default_factory=list)
    blank_columns: list = field(default_factory=list)
    dropped_columns: list = field(default_factory=list)
    # Columns present in the file but NOT in the output sample format - these
    # are removed from the merged output (and reported to the user).
    extra_columns: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    rows: list = field(default_factory=list)
    # Canonical columns of the mode this report was parsed with
    canonical_columns: tuple = POS_CANONICAL_COLUMNS

    @property
    def order_mismatch(self) -> bool:
        """True when the source column order differs from the standard order."""
        return tuple(self.column_order) != self.canonical_columns

    def to_dict(self, status: str = "ok", error: str = "") -> dict:
        return {
            "filename": self.filename,
            "status": status,
            "error": error,
            "sheet": self.sheet,
            "header_row": self.header_row,
            "raw_rows": self.raw_rows,
            "data_rows": self.data_rows,
            "columns_kept": self.columns_kept,
            "column_order": self.column_order,
            "order_mismatch": self.order_mismatch,
            "blank_columns": self.blank_columns,
            "dropped_columns": self.dropped_columns,
            "extra_columns": self.extra_columns,
            "warnings": self.warnings,
        }


def parse_report(file_bytes: bytes, filename: str,
                 mode: ReportMode = POS_DECLINE_MODE) -> FileReport:
    """Parse one report file into canonical rows for the given mode.

    Finds the first sheet with a header row containing "ACQUIRER", skips
    title rows and any repeated header rows, drops blank/spacer columns and
    unrecognized columns, and keeps every value exactly as it is in the
    source file.
    """
    try:
        engine = _detect_engine(file_bytes)
        sheets = _read_grids(file_bytes, engine)
    except ValueError:
        raise
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        raise ValueError(f"'{filename}' is not a readable Excel file ({exc}).") from exc

    report = FileReport(filename=filename)

    for sheet_name, grid in sheets.items():
        if not grid or not any(grid):
            continue

        # pad ragged rows so every row has the same width
        ncols = max(len(row) for row in grid)
        grid = [row + [None] * (ncols - len(row)) for row in grid]

        header_rows = [i for i, row in enumerate(grid) if _is_header_row(row, mode)]
        if not header_rows:
            continue
        header_idx = header_rows[0]
        header_row_set = set(header_rows)

        # Map columns by header NAME (not position), so reshuffled reports
        # still merge into the standard column order. Unknown headers are
        # reported; duplicate headers keep the first occurrence. Columns
        # that exist in the file but NOT in the output sample format (e.g.
        # UTRNNO in a POS export whose sample has no UTRNNO column) are
        # dropped from the output and reported as removed.
        col_map: dict[int, str] = {}
        unknown: list[str] = []
        extra: list[str] = []
        duplicate_warnings: list[str] = []
        for c in range(ncols):
            h = grid[header_idx][c]
            if _is_empty(h):
                continue
            canon = mode.header_aliases.get(str(h).strip().upper())
            if not canon:
                unknown.append(str(h))
                continue
            if canon not in mode.canonical_columns:
                extra.append(canon)
                continue
            if canon in col_map.values():
                duplicate_warnings.append(
                    f"duplicate '{canon}' header in column "
                    f"{get_column_letter(c + 1)} ignored (kept first)"
                )
                continue
            col_map[c] = canon

        # Columns with no header and no data anywhere are blank spacers.
        blank_cols = []
        for c in range(ncols):
            if c in col_map:
                continue
            if not _is_empty(grid[header_idx][c]):
                continue  # header present but unrecognized -> counted above
            has_data = any(
                r not in header_row_set and not _is_empty(grid[r][c])
                for r in range(header_idx + 1, len(grid))
            )
            if not has_data:
                blank_cols.append(get_column_letter(c + 1))

        rows: list[dict] = []
        for r in range(header_idx + 1, len(grid)):
            if r in header_row_set:
                continue  # repeated page-break header
            rec: dict[str, Any] = {}
            any_content = False
            for c, canon in col_map.items():
                v = _cell_value(grid[r][c])
                if v != "":
                    any_content = True
                rec[canon] = v
            if not any_content:
                continue  # fully empty row
            rows.append(rec)

        report.sheet = sheet_name
        report.header_row = header_idx + 1
        report.raw_rows = len(grid)
        report.canonical_columns = mode.canonical_columns
        report.columns_kept = [c for c in mode.canonical_columns if c in set(col_map.values())]
        report.column_order = [col_map[c] for c in sorted(col_map)]
        report.blank_columns = blank_cols
        report.dropped_columns = unknown
        report.extra_columns = sorted(set(extra))
        report.data_rows = len(rows)
        report.rows = rows

        if report.extra_columns:
            report.warnings.append(
                "additional column(s) removed (not in the output sample format): "
                + ", ".join(report.extra_columns)
            )
        missing = [c for c in mode.canonical_columns if c not in set(col_map.values())]
        if missing:
            report.warnings.append(f"missing columns: {', '.join(missing)}")
        if report.order_mismatch:
            report.warnings.append(
                "column order reshuffled - detected: " + ", ".join(report.column_order)
            )
        report.warnings.extend(duplicate_warnings)
        if len(header_rows) > 1:
            report.warnings.append(f"repeated header rows skipped: {len(header_rows) - 1}")
        if len(sheets) > 1:
            others = [n for n, g in sheets.items() if n != sheet_name and g and any(g)]
            if others:
                report.warnings.append(f"other sheets ignored: {', '.join(others)}")
        return report

    markers = ", ".join(sorted(_header_markers(mode)))
    raise ValueError(
        f"No {mode.label} table found in '{filename}': "
        f"no sheet has a header row containing one of: {markers}."
    )


# ---------------------------------------------------------------------------
# Merging
# ---------------------------------------------------------------------------
@dataclass
class MergeResult:
    filename: str
    from_date: str
    to_date: str
    total_rows: int
    per_file: list
    records: list
    resp_counts: dict
    warnings: list
    mode_key: str
    mode_label: str
    workbook_bytes: bytes
    sort_by: str = ""
    sort_dir: str = "asc"
    duplicate_records: list = field(default_factory=list)


def merge_reports(files: list[tuple[str, bytes]], mode_key: str = "pos_decline",
                   skip_workbook: bool = False, sort_by: str = "",
                   sort_dir: str = "asc", dedupe: bool = False) -> MergeResult:
    """Merge any number of (filename, bytes) reports into one workbook.

    ``mode_key`` selects the report type: "pos_decline" (POS decline
    reports), "pos_success" (POS success reports -> SVFE format),
    "pos" (POS transaction reports -> SmartVista daily POS format),
    "atm" (ATM transaction reports) or "qr" (QR transfer-export reports).
    Files that fail to parse are reported per-file and skipped; the merge
    still succeeds as long as at least one file yields transactions.

    ``skip_workbook`` when True skips the expensive workbook build (saves
    significant memory for large files).  The caller can build the workbook
    later via ``build_workbook(result.records, ...)``.

    ``sort_by`` / ``sort_dir`` control the ordering of the merged records.
    ``sort_by`` may be any output column name (e.g. the date or time column,
    an amount, a merchant name, ...) or "date" / "date_time" for the original
    date-then-time ordering. ``sort_dir`` is "asc" or "desc". When ``sort_by``
    is empty, the default date-then-time ordering is used. Values are never
    altered - only an internal smart sort key is used, so dates written with
    spelled months ('15-Aug-26') and numeric dates (20260813) sort together
    correctly.

    ``dedupe`` when True removes fully-duplicate rows - rows whose values are
    identical across every output column are collapsed to a single row (the
    first occurrence is kept). When False (the default) all rows are kept.
    Deduplication is applied *before* sorting, so the surviving row keeps its
    position relative to the other rows. Regardless of ``dedupe`` the result
    also carries ``duplicate_records`` - the rows that appear more than once
    in the merged input (i.e. exactly the rows ``dedupe`` would collapse) -
    so the caller can offer a "download only the duplicates" report.
    """
    mode = MODES.get(mode_key)
    if mode is None:
        raise ValueError(f"Unknown report mode '{mode_key}'.")

    if not files:
        raise ValueError("No files were uploaded.")

    reports: list[FileReport] = []
    per_file: list[dict] = []
    for name, data in files:
        try:
            rep = parse_report(data, name, mode=mode)
            reports.append(rep)
            per_file.append(rep.to_dict(status="ok"))
        except ValueError as exc:
            per_file.append({"filename": name, "status": "error", "error": str(exc)})

    all_rows = [rec for rep in reports for rec in rep.rows]
    if not all_rows:
        raise ValueError(
            f"None of the uploaded files contained parseable {mode.label} transactions."
        )

    # Build the canonical records first (column order fixed, missing values
    # filled with ''), then apply the requested sort. Stored values are NEVER
    # altered - only the internal sort key is normalized, so mixed raw formats
    # (e.g. "115207" vs "01:20:54" or dates spelled as 'Aug 13' vs numbers
    # like 20260813) still order correctly.
    records = [{col: rec.get(col, "") for col in mode.canonical_columns} for rec in all_rows]

    # Build a fingerprint key per record by converting every output-column
    # value to a string.  Empty cells (None, NaN, whitespace, "") all become
    # the empty string so they compare equal regardless of their raw type.
    #
    # Example:
    #   Headers : Class | Age | Grade
    #   Row 1   :  A    |  17 |  5      key -> ("A", "17", "5")
    #   Row 2   :  B    |  17 |  7      key -> ("B", "17", "7")
    #   Row 3   :  A    |  17 |  5      key -> ("A", "17", "5")  <- same as Row 1
    #
    # key_counts after scanning all three rows:
    #   ("A", "17", "5") -> 2   ("B", "17", "7") -> 1
    #
    # duplicate_records -> [Row 1, Row 3]   (both occurrences)
    # After dedupe (if enabled) -> [Row 1, Row 2]   (first occurrence kept)
    #
    # This block always runs so the caller can offer a
    # "download only the duplicates" report even when dedupe is off.
    def _fingerprint(v: Any) -> str:
        """Stable string key for a single cell value.

        Numeric cells are normalized so that 230, 230.0 and 20260804 /
        20260804.0 produce the same key - the same number is often stored as
        an int in one export and as a float in another. String cells (RRN,
        addresses, dates-as-text) are left untouched.
        """
        if _is_empty(v):
            return ""
        if isinstance(v, bool):
            return "1" if v else "0"
        if isinstance(v, (int, float)):
            val = float(v)
            return str(int(val)) if val.is_integer() else repr(v)
        return str(v)

    # Build the list of columns used for the duplicate fingerprint.
    # DUPLICATE_IGNORE_COLUMNS (ACQUIRER, ISSUER, TRANS_TYPE, CURRENCY) are
    # excluded from the check for every mode. If a mode does not have one of
    # those columns it is simply absent from the record dict, so the exclusion
    # has no effect. All remaining columns must match for two rows to be
    # considered duplicates.
    #
    # Example (ATM columns):
    #   Used   : CARD_NUMBER | TRANS_DATE | TRANS_TIME | AMOUNT | RESP |
    #             RRN | UTRNNO | TERMINAL_ID | ADDRESS_NAME
    #   Ignored: ACQUIRER | ISSUER | TRANS_TYPE | CURRENCY
    dup_cols = [
        c for c in mode.canonical_columns
        if c not in DUPLICATE_IGNORE_COLUMNS
    ]

    record_keys = [
        tuple(_fingerprint(rec[c]) for c in dup_cols)
        for rec in records
    ]
    key_counts: dict[tuple, int] = {}
    for k in record_keys:
        key_counts[k] = key_counts.get(k, 0) + 1

    # Rows that appear more than once across the whole merged input.
    # Both the first AND all subsequent occurrences are included so the user
    # can see every instance that was involved in a duplication.
    # (When dedupe=True only the first occurrence survives in ``records``.)
    duplicate_records = [
        rec for rec, k in zip(records, record_keys) if key_counts[k] > 1
    ]

    if dedupe:
        # Collapse fully-duplicate rows (identical values in EVERY column) to
        # one row, keeping the first occurrence. Only the merged output is
        # affected; per-report counts are reported as they were read.
        seen: set = set()
        unique_records: list[dict] = []
        for rec, k in zip(records, record_keys):
            if k in seen:
                continue
            seen.add(k)
            unique_records.append(rec)
        records = unique_records

    if sort_by in ("", "date_time", "date"):
        # default: sort by date (spelled months or numbers), then time
        records.sort(
            key=lambda r: (
                smart_sort_key(r.get(mode.date_column)),
                _normalize_time(r.get(mode.time_column)),
            )
        )
        if sort_dir == "desc":
            _reverse_date_time_sort(records, mode)
    elif sort_by in mode.canonical_columns:
        records = sort_records(records, sort_by, descending=(sort_dir == "desc"))
    # unknown sort column -> leave in source order

    dates = sorted(
        {
            _normalize_date(r[mode.date_column])
            for r in records
            if not _is_empty(r[mode.date_column])
        }
    )
    from_date = dates[0] if dates else ""
    to_date = dates[-1] if dates else ""

    resp_counts = dict(
        Counter(
            _count_key(r[mode.resp_column])
            for r in records
            if not _is_empty(r[mode.resp_column])
        )
    )

    # ---- column-consistency check across reports -------------------------
    # Every uploaded file should have at least the columns of the output
    # sample format. If a column is missing, warn explicitly, naming the
    # missing column(s) and the file they are missing from.
    warnings: list[str] = []

    col_files: dict[str, set[str]] = {}
    for rep in reports:
        for col in rep.columns_kept:
            col_files.setdefault(col, set()).add(rep.filename)

    counts = {rep.filename: len(rep.columns_kept) for rep in reports}
    if len(set(counts.values())) > 1:
        warnings.append(
            "Unbalanced column counts between reports: "
            + ", ".join(f"'{name}' has {count}" for name, count in sorted(counts.items()))
        )

    for rep in reports:
        missing_all = [
            c for c in mode.canonical_columns if c not in rep.columns_kept
        ]
        if not missing_all:
            continue
        in_others = [c for c in missing_all if c in col_files]
        sample_only = [c for c in missing_all if c not in col_files]
        if in_others:
            msg = (
                f"Missing column(s) in '{rep.filename}': {', '.join(in_others)}"
                " - these columns are present in other reports"
            )
            rep.warnings.append(msg)
            warnings.append(msg)
        if sample_only:
            msg = (
                f"Missing column(s) in '{rep.filename}': {', '.join(sample_only)}"
                f" - these columns are in the {mode.sample_label} output format"
            )
            rep.warnings.append(msg)
            warnings.append(msg)

    if mode.always_show_range or from_date != to_date:
        date_part = f"{mode.file_date(from_date)}_to_{mode.file_date(to_date)}"
    else:
        date_part = mode.file_date(from_date)
    filename = f"{mode.output_prefix}_{date_part}_Merged.xlsx"

    return MergeResult(
        filename=filename,
        from_date=from_date,
        to_date=to_date,
        total_rows=len(records),
        per_file=per_file,
        records=records,
        duplicate_records=duplicate_records,
        resp_counts=resp_counts,
        warnings=warnings,
        mode_key=mode.key,
        mode_label=mode.label,
        workbook_bytes=b"" if skip_workbook else build_workbook(records, from_date, to_date, mode),
        sort_by=sort_by,
        sort_dir=sort_dir,
    )


# ---------------------------------------------------------------------------
# Output workbook (POS_Transaction_Decline_Report layout, blanks removed)
# ---------------------------------------------------------------------------
def _as_cell(value: Any) -> Any:
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return value


def build_workbook(records: list[dict], from_date: str = "", to_date: str = "",
                   mode: ReportMode = POS_DECLINE_MODE) -> bytes:
    """Build the merged workbook in memory and return its bytes.

    POS decline mode uses the POS_Transaction_Decline_Report layout (title
    block + header row). POS and ATM modes use the SmartVista daily layout
    (header in row 1) and write the output with a fast manual XML writer so
    the very large ATM reports build quickly.
    """
    if mode.title_rows == 0:
        return _build_workbook_streaming(records, mode)
    return _build_workbook_normal(records, from_date, to_date, mode)


def _build_workbook_normal(records: list[dict], from_date: str, to_date: str,
                           mode: ReportMode) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = mode.sheet_name

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    label_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="1F2937")
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    # --- title block (POS layout) ---
    last_col = get_column_letter(len(mode.canonical_columns))
    a1 = ws.cell(row=1, column=1, value="Report name:")
    a1.font = label_font
    ws.merge_cells(f"B1:{last_col}1")
    b1 = ws.cell(row=1, column=2, value=mode.report_title)
    b1.font = title_font
    b1.alignment = Alignment(horizontal="left", vertical="center")

    for row, label, value in ((2, "From Date:", from_date), (3, "To Date:", to_date)):
        c = ws.cell(row=row, column=1, value=label)
        c.font = label_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        v = ws.cell(row=row, column=2, value=_as_cell(value))
        v.font = label_font
        if isinstance(v.value, int):
            v.number_format = "0"

    # --- header row ---
    header_row = mode.title_rows + 1  # 4 for POS
    for c, name in enumerate(mode.canonical_columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # --- data rows ---
    numeric_fmt_cols = set(mode.numeric_fmt_cols)
    center_cols = {mode.date_column, mode.time_column, mode.resp_column, "Reversal"}
    for r, rec in enumerate(records, start=header_row + 1):
        for c, col in enumerate(mode.canonical_columns, start=1):
            v = rec.get(col, "")
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                # Display-only guard: show full digits instead of scientific
                # notation. The stored value itself is never changed.
                if col in numeric_fmt_cols:
                    cell.number_format = "0"
            if col in center_cols:
                cell.alignment = center

    # --- layout ---
    for col, w in mode.column_widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_workbook_streaming(records: list[dict], mode: ReportMode) -> bytes:
    """Build the SmartVista-daily-style workbook (header in row 1, plain
    sheet - no fills, borders, column widths or frozen panes, exactly like
    the sample output reports). Uses a fast manual XML writer; falls back to
    openpyxl's write-only mode if anything unexpected happens."""
    try:
        return _build_workbook_manual_xml(records, mode)
    except Exception:  # noqa: BLE001 - fall back to the battle-tested writer
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=mode.sheet_name)
        ws.append(list(mode.canonical_columns))
        for rec in records:
            ws.append([rec.get(col, "") for col in mode.canonical_columns])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def _col_letters(n: int) -> list[str]:
    """Column letters for the first n columns (A..Z, AA, AB, ...)."""
    letters = []
    for i in range(1, n + 1):
        s = ""
        while i:
            i, r = divmod(i - 1, 26)
            s = chr(65 + r) + s
        letters.append(s)
    return letters


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;").replace("<", "&lt;")
        .replace(">", "&gt;").replace('"', "&quot;")
    )


def _xml_number(value: Any) -> str:
    """Serialize a number exactly as stored (int stays int, floats keep
    their decimal/scientific representation)."""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    s = repr(value)
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace("e", "E")


# ---------------------------------------------------------------------------
# Filtered multi-sheet workbook
# ---------------------------------------------------------------------------
def build_filtered_workbook(
    records: list[dict],
    columns: list[str],
    sheet_defs: list[dict],
    mode: ReportMode,
) -> bytes:
    """Build a workbook where each sheet is a filtered subset of *records*.

    ``sheet_defs`` is a list of dicts, each with:
        {"name": "sheet_name", "filters": {"ISSUER": ["Abay", "CBE"], "RESP_CODE": ["00"]}}

    Filter values can be single strings or lists.  Within a column: OR.
    Between columns: AND.  Empty filters dict = all rows.
    means "all rows" (unfiltered).  Sheet names are truncated to 31 chars
    (Excel limit) and sanitised.
    """
    wb = Workbook()
    # Remove the default sheet — we'll create one per definition
    wb.remove(wb.active)

    if mode.title_rows == 0:
        # SmartVista-style: plain header row, no title block
        for sd in sheet_defs:
            ws = wb.create_sheet(title=_safe_sheet_name(sd["name"]))
            _write_filtered_sheet_streaming(ws, records, columns, sd.get("filters", {}))
    else:
        # POS-style: title block + header row
        for sd in sheet_defs:
            ws = wb.create_sheet(title=_safe_sheet_name(sd["name"]))
            _write_filtered_sheet_normal(ws, records, columns, sd.get("filters", {}), mode)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_sheet_name(name: str) -> str:
    """Sanitise a sheet name for Excel (max 31 chars, no special chars)."""
    bad = r'[\\/*?\[\]:]'
    name = re.sub(bad, "_", name)
    return name[:31] or "Sheet"


def _write_filtered_sheet_normal(ws, records, columns, filters, mode):
    """Write a POS-style filtered sheet with title block."""
    filtered = _apply_filters(records, filters)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    label_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="1F2937")
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    last_col = get_column_letter(len(columns))
    ws.cell(row=1, column=1, value="Report name:").font = label_font
    ws.merge_cells(f"B1:{last_col}1")
    b1 = ws.cell(row=1, column=2, value=mode.report_title)
    b1.font = title_font
    b1.alignment = Alignment(horizontal="left", vertical="center")

    filter_desc = ", ".join(
        f"{k}={','.join(v) if isinstance(v, (list,tuple,set)) else v}"
        for k, v in filters.items()
    ) if filters else "All"
    ws.cell(row=2, column=1, value="Filter:").font = label_font
    ws.merge_cells("B2:C2")
    ws.cell(row=2, column=2, value=filter_desc).font = label_font

    ws.cell(row=3, column=1, value="Rows:").font = label_font
    ws.cell(row=3, column=2, value=str(len(filtered))).font = label_font

    header_row = mode.title_rows + 1
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    numeric_fmt_cols = set(mode.numeric_fmt_cols)
    center_cols = {mode.date_column, mode.time_column, mode.resp_column, "Reversal"}
    for r, rec in enumerate(filtered, start=header_row + 1):
        for c, col in enumerate(columns, start=1):
            v = rec.get(col, "")
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if col in numeric_fmt_cols:
                    cell.number_format = "0"
            if col in center_cols:
                cell.alignment = center

    for col_letter, w in mode.column_widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False


def _write_filtered_sheet_streaming(ws, records, columns, filters):
    """Write a SmartVista-style filtered sheet (plain header row)."""
    filtered = _apply_filters(records, filters)
    ws.append(columns)
    for rec in filtered:
        ws.append([rec.get(col, "") for col in columns])


def _apply_filters(records: list[dict], filters: dict) -> list[dict]:
    """Filter records by exact match (case-insensitive) on each filter column.

    Each filter value can be a single string or a list of strings.
    Within a column the logic is OR (match any of the values).
    Between columns the logic is AND (match every column's filter).
    """
    if not filters:
        return records
    result = records
    for col, vals in filters.items():
        if isinstance(vals, (list, tuple, set)):
            allowed = {str(v).strip().upper() for v in vals if v}
        else:
            allowed = {str(vals).strip().upper()}
        if not allowed:
            continue
        result = [
            r for r in result
            if str(r.get(col, "")).strip().upper() in allowed
        ]
    return result


def _build_workbook_manual_xml(records: list[dict], mode: ReportMode) -> bytes:
    """Write the workbook by building the OOXML parts directly (no styling),
    which is several times faster than any general-purpose writer for the
    very large reports. Cell values are written exactly as they are stored."""
    cols = list(mode.canonical_columns)
    letters = _col_letters(len(cols))

    # ---- sheet data -------------------------------------------------------
    row_xml = ["<row r=\"1\">" + "".join(
        f'<c r="{letters[c]}1" t="inlineStr"><is><t>{_xml_escape(h)}</t></is></c>'
        for c, h in enumerate(cols)
    ) + "</row>"]
    append_row = row_xml.append
    for rn, rec in enumerate(records, start=2):
        cells = []
        for c, col in enumerate(cols):
            v = rec.get(col, "")
            ref = f"{letters[c]}{rn}"
            if isinstance(v, str):
                cells.append(
                    f'<c r="{ref}" t="inlineStr"><is><t>{_xml_escape(v)}</t></is></c>'
                )
            elif isinstance(v, bool):
                cells.append(f'<c r="{ref}" t="b"><v>{"1" if v else "0"}</v></c>')
            elif v is None or v == "":
                cells.append(f'<c r="{ref}"/>')
            else:
                cells.append(f'<c r="{ref}"><v>{_xml_number(v)}</v></c>')
        append_row(f"<row r=\"{rn}\">" + "".join(cells) + "</row>")
    sheet_data = "".join(row_xml)

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f"<sheetData>{sheet_data}</sheetData>"
        "</worksheet>"
    )

    # ---- package parts ----------------------------------------------------
    sheet_name = _xml_escape(mode.sheet_name)
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        "</Types>"
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets><sheet name="{sheet_name}" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        "</Relationships>"
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/><scheme val="minor"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )
    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>Report Merger</Application>'
        "</Properties>"
    )
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "</cp:coreProperties>"
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", root_rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        zf.writestr("xl/styles.xml", styles_xml)
        zf.writestr("docProps/app.xml", app_xml)
        zf.writestr("docProps/core.xml", core_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buf.getvalue()
