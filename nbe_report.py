"""NBE Institution Report generator for POS (Daily) and ATM (Daily) merged reports.

Normalises ISSUER and ACQUIRER bank names to canonical NBE institution names,
filters by transaction type (POS purchase vs ATM cash withdrawal) and response code (-1, -1.0),
aggregates transaction count and total monetary value, and outputs structured DataFrames
and styled Excel workbooks matching NBE REPORT layout.
"""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Standard NBE Institution Order (31 Institutions)
# ---------------------------------------------------------------------------
STANDARD_NBE_BANKS = [
    "Abay Bank",
    "Addis Bank",
    "Ahadu Bank",
    "Amhara Bank",
    "Awash Bank",
    "Birhan Bank",
    "BOA",
    "Bunna Bank",
    "CBE",
    "CBO",
    "Global Bank",
    "Dashen Bank",
    "Enat Bank",
    "Gadda Bank",
    "Goh Betoch Bank",
    "Hijra Bank",
    "Lion Bank",
    "Nib Bank",
    "Oromia Bank",
    "Rammis Bank",
    "Santim Pay",
    "Sinqee Bank",
    "Sidama Bank",
    "Siket Bank",
    "Tseday Bank",
    "Tsehay Bank",
    "United Bank",
    "Wegagen Bank",
    "Yagout Pay",
    "Zamzam Bank",
    "Zemen Bank",
]

# ---------------------------------------------------------------------------
# Comprehensive Bank Name Alias Mapping
# ---------------------------------------------------------------------------
_NBE_ALIAS_MAP: dict[str, str] = {
    # Abay
    "abay": "Abay Bank",
    "abay bank": "Abay Bank",
    # Addis
    "addis": "Addis Bank",
    "addis bank": "Addis Bank",
    "addis int": "Addis Bank",
    "addis int bank": "Addis Bank",
    "addis international": "Addis Bank",
    "addis international bank": "Addis Bank",
    # Ahadu
    "ahadu": "Ahadu Bank",
    "ahadu bank": "Ahadu Bank",
    "ahadu ebirr": "Ahadu Bank",
    # Amhara
    "amhara": "Amhara Bank",
    "amhara bank": "Amhara Bank",
    "amharaethbirr": "Amhara Bank",
    # Awash
    "awash": "Awash Bank",
    "awash bank": "Awash Bank",
    "aib": "Awash Bank",
    # Birhan / Berhan
    "birhan": "Birhan Bank",
    "birhan bank": "Birhan Bank",
    "berhan": "Birhan Bank",
    "berhan bank": "Birhan Bank",
    # Bank of Abyssinia / BOA
    "boa": "BOA",
    "abyssinia": "BOA",
    "abyssinia bank": "BOA",
    "bank of abyssinia": "BOA",
    # Bunna
    "bunna": "Bunna Bank",
    "bunna bank": "Bunna Bank",
    "buna": "Bunna Bank",
    "buna bank": "Bunna Bank",
    "bunna int": "Bunna Bank",
    "bunna int bank": "Bunna Bank",
    "bunna international bank": "Bunna Bank",
    # CBE
    "cbe": "CBE",
    "commercial bank": "CBE",
    "commercial bank of ethiopia": "CBE",
    "cbé": "CBE",
    "cbébirr": "CBE",
    "cbébírr": "CBE",
    # CBO / Coop
    "cbo": "CBO",
    "cbo switch": "CBO",
    "coop": "CBO",
    "coop bank": "CBO",
    "coop bank of oromia": "CBO",
    "cooperative bank of oromia": "CBO",
    "coopay-e-birr": "CBO",
    "coopay": "CBO",
    # Global / Debub
    "global": "Global Bank",
    "global bank": "Global Bank",
    "debub": "Global Bank",
    "debub bank": "Global Bank",
    "dedebit": "Global Bank",
    # Dashen
    "dashen": "Dashen Bank",
    "dashen bank": "Dashen Bank",
    "db": "Dashen Bank",
    # Enat
    "enat": "Enat Bank",
    "enat bank": "Enat Bank",
    # Gadda / Gadaa
    "gadda": "Gadda Bank",
    "gadda bank": "Gadda Bank",
    "gadaa": "Gadda Bank",
    "gadaa bank": "Gadda Bank",
    "gada": "Gadda Bank",
    "gedaa": "Gadda Bank",
    "gedaa bank": "Gadda Bank",
    # Goh Betoch
    "goh": "Goh Betoch Bank",
    "goh betoch": "Goh Betoch Bank",
    "goh betoch bank": "Goh Betoch Bank",
    # Hijra
    "hijra": "Hijra Bank",
    "hijra bank": "Hijra Bank",
    # Lion
    "lion": "Lion Bank",
    "lion bank": "Lion Bank",
    "lion int": "Lion Bank",
    "lion int bank": "Lion Bank",
    "lion international bank": "Lion Bank",
    "lib": "Lion Bank",
    # Nib
    "nib": "Nib Bank",
    "nib bank": "Nib Bank",
    "nib int": "Nib Bank",
    "nib int bank": "Nib Bank",
    "nib international": "Nib Bank",
    "nib international bank": "Nib Bank",
    "nibbirr": "Nib Bank",
    # Oromia
    "oromia": "Oromia Bank",
    "oromia bank": "Oromia Bank",
    "oib": "Oromia Bank",
    # Rammis
    "rammis": "Rammis Bank",
    "rammis bank": "Rammis Bank",
    "raamis": "Rammis Bank",
    "raammis bank": "Rammis Bank",
    "ramis": "Rammis Bank",
    "ramis bank": "Rammis Bank",
    # Santim Pay
    "santimpay": "Santim Pay",
    "santim pay": "Santim Pay",
    "santim": "Santim Pay",
    # Sinqee
    "sinqee": "Sinqee Bank",
    "sinqee bank": "Sinqee Bank",
    "siinqee": "Sinqee Bank",
    "siinqee bank": "Sinqee Bank",
    "siinqee wallet": "Sinqee Bank",
    # Sidama
    "sidama": "Sidama Bank",
    "sidama bank": "Sidama Bank",
    # Siket
    "siket": "Siket Bank",
    "siket bank": "Siket Bank",
    # Tseday
    "tseday": "Tseday Bank",
    "tseday bank": "Tseday Bank",
    "tsedey": "Tseday Bank",
    "tsedey bank": "Tseday Bank",
    # Tsehay
    "tsehay": "Tsehay Bank",
    "tsehay bank": "Tsehay Bank",
    # United / Hibret
    "united": "United Bank",
    "united bank": "United Bank",
    "hibret": "United Bank",
    "hibret bank": "United Bank",
    "ub": "United Bank",
    "h-cash": "United Bank",
    # Wegagen
    "wegagen": "Wegagen Bank",
    "wegagen bank": "Wegagen Bank",
    "wb": "Wegagen Bank",
    "wegagen e-birr": "Wegagen Bank",
    # Yagout Pay
    "yagoutpay": "Yagout Pay",
    "yagout pay": "Yagout Pay",
    "yagout": "Yagout Pay",
    # Zamzam
    "zamzam": "Zamzam Bank",
    "zamzam bank": "Zamzam Bank",
    "zam zam": "Zamzam Bank",
    "zam zam bank": "Zamzam Bank",
    # Zemen
    "zemen": "Zemen Bank",
    "zemen bank": "Zemen Bank",
    "zb": "Zemen Bank",
}


def normalize_nbe_bank(name: Any) -> str:
    """Return the canonical NBE display name for a bank, or title-cased cleaned value."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ""
    s = str(name).strip()
    if not s:
        return ""
    key = s.lower()
    if key in _NBE_ALIAS_MAP:
        return _NBE_ALIAS_MAP[key]

    # Try removing trailing suffixes
    for suffix in (" bank", " int bank", " international bank", " ebirr", " e-birr", " wallet"):
        if key.endswith(suffix):
            base_key = key[:-len(suffix)].strip()
            if base_key in _NBE_ALIAS_MAP:
                return _NBE_ALIAS_MAP[base_key]

    return s.title()


def _is_success_resp(val: Any) -> bool:
    """Return True if response code is -1 or -1.0."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    s = str(val).strip()
    if s in ("-1", "-1.0"):
        return True
    try:
        return float(s) == -1.0
    except (ValueError, TypeError):
        return False


def generate_nbe_report(records: list[dict[str, Any]], mode_key: str) -> pd.DataFrame:
    """Generate NBE Institution Summary DataFrame for POS or ATM daily records.

    Filters:
    - POS: TRANS_TYPE in ('pos purchase', 'purchase'), RESP in (-1, -1.0)
    - ATM: TRANS_TYPE in ('atm cash withdrawal', 'cash withdrawal'), RESP in (-1, -1.0)
    """
    if mode_key == "pos":
        valid_types = {"pos purchase", "purchase"}
        trans_label = "PURCHASE"
    elif mode_key == "atm":
        valid_types = {"atm cash withdrawal", "cash withdrawal"}
        trans_label = "CASH WITHDRAWAL"
    else:
        raise ValueError(f"NBE report is only supported for 'pos' and 'atm' modes, got '{mode_key}'")

    # Data aggregators per institution
    stats: dict[str, dict[str, float]] = {}

    def _get_bank_stat(b: str) -> dict[str, float]:
        if b not in stats:
            stats[b] = {
                "issuer_count": 0,
                "issuer_amount": 0.0,
                "acquirer_count": 0,
                "acquirer_amount": 0.0,
            }
        return stats[b]

    # Process matching records
    for r in records:
        # Check transaction type
        t_type = str(r.get("TRANS_TYPE", "")).strip().lower()
        if t_type not in valid_types:
            continue

        # Check response code (-1 / -1.0)
        resp_val = r.get("RESP")
        if resp_val is None and "RESP_CODE" in r:
            resp_val = r.get("RESP_CODE")
        if resp_val is None and "RESP CODE" in r:
            resp_val = r.get("RESP CODE")
        if resp_val is None and "STATUS" in r:
            resp_val = r.get("STATUS")

        if not _is_success_resp(resp_val):
            continue

        # Amount
        try:
            amt = float(r.get("AMOUNT", 0) or 0)
        except (ValueError, TypeError):
            amt = 0.0

        iss = normalize_nbe_bank(r.get("ISSUER"))
        acq = normalize_nbe_bank(r.get("ACQUIRER"))

        if iss:
            st_iss = _get_bank_stat(iss)
            st_iss["issuer_count"] += 1
            st_iss["issuer_amount"] += amt

        if acq:
            st_acq = _get_bank_stat(acq)
            st_acq["acquirer_count"] += 1
            st_acq["acquirer_amount"] += amt

    # Build full institution list: Standard 31 NBE banks first, followed by any extras
    bank_list = list(STANDARD_NBE_BANKS)
    seen_banks = set(bank_list)
    extra_banks = sorted([b for b in stats.keys() if b and b not in seen_banks])
    bank_list.extend(extra_banks)

    rows = []
    tot_iss_cnt = 0
    tot_iss_amt = 0.0
    tot_acq_cnt = 0
    tot_acq_amt = 0.0

    for idx, b in enumerate(bank_list, start=1):
        s = stats.get(b, {"issuer_count": 0, "issuer_amount": 0.0, "acquirer_count": 0, "acquirer_amount": 0.0})
        i_cnt = int(s["issuer_count"])
        i_amt = round(float(s["issuer_amount"]), 2)
        a_cnt = int(s["acquirer_count"])
        a_amt = round(float(s["acquirer_amount"]), 2)

        tot_iss_cnt += i_cnt
        tot_iss_amt += i_amt
        tot_acq_cnt += a_cnt
        tot_acq_amt += a_amt

        rows.append({
            "S/N": idx,
            "BANKS": b,
            f"{trans_label} As Issuer (Count)": i_cnt,
            f"{trans_label} As Issuer (Amount ETB)": i_amt,
            f"{trans_label} As Acquirer (Count)": a_cnt,
            f"{trans_label} As Acquirer (Amount ETB)": a_amt,
        })

    # Summary row
    rows.append({
        "S/N": "",
        "BANKS": "Total",
        f"{trans_label} As Issuer (Count)": tot_iss_cnt,
        f"{trans_label} As Issuer (Amount ETB)": round(tot_iss_amt, 2),
        f"{trans_label} As Acquirer (Count)": tot_acq_cnt,
        f"{trans_label} As Acquirer (Amount ETB)": round(tot_acq_amt, 2),
    })

    return pd.DataFrame(rows)


def build_nbe_report_excel(df: pd.DataFrame, mode_key: str) -> bytes:
    """Build formatted Excel file matching NBE Report layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NBE Report"

    trans_label = "PURCHASE" if mode_key == "pos" else "CASH WITHDRAWAL"

    # Styling definitions
    font_family = "Arial"

    header_title_font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    header_title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    sub_header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    sub_header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    col_hdr_font = Font(name=font_family, size=10, bold=True, color="1F4E78")
    col_hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    data_font = Font(name=font_family, size=10)
    total_font = Font(name=font_family, size=11, bold=True)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    thick_top_double_bottom = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )

    # Row 1: Report Main Header
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"NBE REPORT - {mode_key.upper()} ({trans_label})"
    title_cell.font = header_title_font
    title_cell.fill = header_title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2: Section Header (BANKS | As Issuer | As Acquirer)
    ws["A2"] = "S/N"
    ws["B2"] = "BANKS"
    ws.merge_cells("C2:D2")
    ws["C2"] = f"{trans_label} As Issuer"
    ws.merge_cells("E2:F2")
    ws["E2"] = f"{trans_label} As Acquirer"

    for col in range(1, 7):
        cell = ws.cell(row=2, column=col)
        cell.font = sub_header_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # Row 3: Sub-columns (S/N | BANKS | Count | Amount | Count | Amount)
    ws["A3"] = "S/N"
    ws["B3"] = "BANKS"
    ws["C3"] = "Count"
    ws["D3"] = "Amount (ETB)"
    ws["E3"] = "Count"
    ws["F3"] = "Amount (ETB)"

    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.font = col_hdr_font
        cell.fill = col_hdr_fill
        cell.alignment = Alignment(horizontal="center" if col in (1, 3, 5) else ("left" if col == 2 else "right"), vertical="center")
    ws.row_dimensions[3].height = 20

    # Data Rows
    current_row = 4
    total_row_idx = len(df) + 3

    for idx, row in df.iterrows():
        is_total = (idx == len(df) - 1)
        row_num = current_row

        ws.cell(row=row_num, column=1, value=row.iloc[0])
        ws.cell(row=row_num, column=2, value=row.iloc[1])

        c_cnt = ws.cell(row=row_num, column=3, value=row.iloc[2])
        c_amt = ws.cell(row=row_num, column=4, value=row.iloc[3])
        a_cnt = ws.cell(row=row_num, column=5, value=row.iloc[4])
        a_amt = ws.cell(row=row_num, column=6, value=row.iloc[5])

        # Number formats
        c_cnt.number_format = "#,##0"
        c_amt.number_format = "#,##0.00"
        a_cnt.number_format = "#,##0"
        a_amt.number_format = "#,##0.00"

        for col in range(1, 7):
            cell = ws.cell(row=row_num, column=col)
            if is_total:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thick_top_double_bottom
            else:
                cell.font = data_font
                cell.border = thin_border

            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col == 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")

        ws.row_dimensions[row_num].height = 20
        current_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["F"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
