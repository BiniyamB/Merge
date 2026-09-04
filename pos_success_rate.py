"""POS Success Rate Report generator (Issuer View).

Groups POS transactions by ISSUER institution, computes decline response code breakdowns,
calculates Cardholder-Related Declines (RC 821, 901, 904, 906, 911, 912, 914, 915),
Total Success, Total Attempted Transactions, and Adjusted Success Rate %, and exports
formatted DataFrames and styled Excel workbooks with native Excel formulas matching POS SUCCESS RATE layout.
"""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nbe_report import STANDARD_NBE_BANKS, normalize_nbe_bank

# ---------------------------------------------------------------------------
# Cardholder Related Declines Response Codes
# ---------------------------------------------------------------------------
CARDHOLDER_RC = {"821", "901", "904", "906", "911", "912", "914", "915"}

# ---------------------------------------------------------------------------
# Official Response Code Descriptions & Remarks Lookup
# ---------------------------------------------------------------------------
RESPONSE_CODE_LOOKUP: list[dict[str, str]] = [
    {"rc": "503", "description": "Not valid EMV transaction", "remark": "Not valid EMV transaction"},
    {"rc": "504", "description": "Card Operating rule does Not allow this transaction", "remark": "Card Operating rule does Not allow this transaction"},
    {"rc": "801", "description": "Time out", "remark": "Bank Core banking system (CBS) or Issuer Switch did not respond"},
    {"rc": "802", "description": "Issuer not operative", "remark": "Bank is disconnected from Ethswitch"},
    {"rc": "804", "description": "Card is not permitted", "remark": "Card is not permitted by the system for transaction"},
    {"rc": "805", "description": "ERROR - 805", "remark": ""},
    {"rc": "812", "description": "Message received was in wrong format", "remark": "The message was received in wrong format which can not be parsable by the system"},
    {"rc": "821", "description": "Wrong PIN, Excessive PIN Failures", "remark": "Wrong PIN is entered, wrong PIN is entered 3 and more times"},
    {"rc": "827", "description": "Do not honor transaction", "remark": "Generic Response from the Bank (Exactly not known)"},
    {"rc": "857", "description": "Requested amount was out of range allowed by the issuer.", "remark": "Requested amount was out of range allowed by the issuer."},
    {"rc": "858", "description": "Processing error during MAC-related HSM command", "remark": "Error related with Keys"},
    {"rc": "862", "description": "Excessive PIN failures, do not capture", "remark": "Wrong PIN is entered 3 times & card is not captured by ATM"},
    {"rc": "873", "description": "Issuing BIN is unknown", "remark": "Card is unknown by Ethswitch"},
    {"rc": "878", "description": "Account is locked", "remark": "Account is locked in CBS"},
    {"rc": "886", "description": "Card inactive", "remark": "Card is not in active status"},
    {"rc": "901", "description": "Invalid PIN", "remark": "Customer enter invalid PIN or wrong PIN"},
    {"rc": "902", "description": "Cannot Process Transaction", "remark": "The transaction is cannot be processed by the system due to some format error"},
    {"rc": "904", "description": "Excessive PIN failures, capture", "remark": "Wrong PIN is entered 3 times & card is captured by ATM"},
    {"rc": "905", "description": "Invalid Card", "remark": "Card is not found in Data base"},
    {"rc": "906", "description": "Card Has Expired", "remark": "Card Has Expired"},
    {"rc": "909", "description": "Invalid card, capture.", "remark": "The card is not valid or cannot be used for transaction and captured by the ATM"},
    {"rc": "911", "description": "Withdrawal Limit Reached - Retry", "remark": "Maximum limit of amount a customer can withdraw is reached"},
    {"rc": "912", "description": "Withdrawal Limit Exceeded", "remark": "Maximum limit of amount a customer can withdraw is exceeded"},
    {"rc": "913", "description": "Transaction Type Not Supported By Institution", "remark": "Transaction Type Not Supported By Institution"},
    {"rc": "914", "description": "Invalid Account", "remark": "wrong Account linked to card"},
    {"rc": "915", "description": "Insufficient Funds", "remark": "Customer wants to withdraw cash more than what he has in the account"},
    {"rc": "917", "description": "ATM or POS limit exceeded", "remark": "ATM or POS transaction amount limit exceeded"},
    {"rc": "939", "description": "No such response code from network", "remark": "Unknown response code responded by the Bank"},
    {"rc": "952", "description": "Fraud is suspected", "remark": "This Response code is sent when transaction is done using fallback method"},
    {"rc": "959", "description": "System malfunction", "remark": "System malfunction"},
    {"rc": "979", "description": "Invalid account type", "remark": "Customer has savings account but select checking account while performing transaction or vice versa"},
    {"rc": "988", "description": "Service not available at that time", "remark": "Service not available at the time when transaction is performed"},
]

_LOOKUP_MAP = {item["rc"]: item for item in RESPONSE_CODE_LOOKUP}


def parse_rc(val: Any) -> str:
    """Normalize response code into standard string key."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    if s in ("-1", "-1.0", "00", "0"):
        return "-1"
    try:
        f = float(s)
        if f in (-1.0, 0.0):
            return "-1"
        if f.is_integer():
            return str(int(f))
        return s
    except (ValueError, TypeError):
        return s


def get_rate_style(rate: float) -> tuple[PatternFill, Font]:
    """Return PatternFill and Font according to Success Rate % tier rules:
    - 97% - 100%: Green
    - 86% - 96%: Yellow
    - 79% - 85%: Light Yellow
    - <= 78%: Red
    """
    if rate >= 0.97:
        return (
            PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            Font(name="Arial", size=10, bold=True, color="006100"),
        )
    elif rate >= 0.86:
        return (
            PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            Font(name="Arial", size=10, bold=True, color="9C6500"),
        )
    elif rate >= 0.79:
        return (
            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            Font(name="Arial", size=10, bold=True, color="7F6000"),
        )
    else:
        return (
            PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            Font(name="Arial", size=10, bold=True, color="9C0006"),
        )


def generate_pos_success_rate_report(records: list[dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Generate POS Success Rate Matrix DataFrame (grouped strictly by ISSUER)
    and Response Code Description DataFrame.
    """
    if not records:
        empty_matrix = pd.DataFrame(columns=["RC/BANK NAME", "Total"])
        desc_df = pd.DataFrame(RESPONSE_CODE_LOOKUP).rename(
            columns={"rc": "Response Code", "description": "Description", "remark": "Remark"}
        )
        return empty_matrix, desc_df

    parsed_rows = []
    for r in records:
        iss = normalize_nbe_bank(r.get("ISSUER"))
        if not iss:
            continue

        resp_val = r.get("RESP")
        if resp_val is None and "RESP_CODE" in r:
            resp_val = r.get("RESP_CODE")
        if resp_val is None and "RESP CODE" in r:
            resp_val = r.get("RESP CODE")
        if resp_val is None and "STATUS" in r:
            resp_val = r.get("STATUS")

        rc = parse_rc(resp_val)
        parsed_rows.append({"issuer": iss, "rc": rc})

    df_parsed = pd.DataFrame(parsed_rows)
    if df_parsed.empty:
        empty_matrix = pd.DataFrame(columns=["RC/BANK NAME", "Total"])
        desc_df = pd.DataFrame(RESPONSE_CODE_LOOKUP).rename(
            columns={"rc": "Response Code", "description": "Description", "remark": "Remark"}
        )
        return empty_matrix, desc_df

    present_issuers = set(df_parsed["issuer"].unique())
    issuers = [b for b in STANDARD_NBE_BANKS if b in present_issuers]
    extra_issuers = sorted([b for b in present_issuers if b not in set(STANDARD_NBE_BANKS)])
    issuers.extend(extra_issuers)

    all_rcs = set(df_parsed["rc"].unique())
    decline_rcs = sorted([r for r in all_rcs if r and r != "-1"], key=lambda x: int(x) if x.isdigit() else 9999)

    matrix_counts: dict[str, dict[str, int]] = {rc: {b: 0 for b in issuers} for rc in decline_rcs}
    success_counts: dict[str, int] = {b: 0 for b in issuers}

    for _, row in df_parsed.iterrows():
        iss = row["issuer"]
        rc = row["rc"]
        if iss not in issuers:
            continue
        if rc in ("-1", ""):
            success_counts[iss] += 1
        elif rc in matrix_counts:
            matrix_counts[rc][iss] += 1

    rows = []

    # 1. Individual Response Code Rows
    for rc in decline_rcs:
        r_data: dict[str, Any] = {"RC/BANK NAME": str(rc)}
        tot = 0
        for b in issuers:
            cnt = matrix_counts[rc][b]
            r_data[b] = cnt
            tot += cnt
        r_data["Total"] = tot
        rows.append(r_data)

    # 2. Total Decline Row
    tot_decline_row: dict[str, Any] = {"RC/BANK NAME": "Total Decline"}
    tot_decline_all = 0
    for b in issuers:
        b_dec = sum(matrix_counts[rc][b] for rc in decline_rcs)
        tot_decline_row[b] = b_dec
        tot_decline_all += b_dec
    tot_decline_row["Total"] = tot_decline_all
    rows.append(tot_decline_row)

    # 3. Successful Pos T Row
    succ_row: dict[str, Any] = {"RC/BANK NAME": "Successful Pos T"}
    tot_succ_pos_all = 0
    for b in issuers:
        cnt = success_counts[b]
        succ_row[b] = cnt
        tot_succ_pos_all += cnt
    succ_row["Total"] = tot_succ_pos_all
    rows.append(succ_row)

    # 4. card holder rel dec Row
    ch_dec_row: dict[str, Any] = {"RC/BANK NAME": "card holder rel dec"}
    tot_ch_dec_all = 0
    for b in issuers:
        ch_cnt = sum(matrix_counts[rc][b] for rc in decline_rcs if rc in CARDHOLDER_RC)
        ch_dec_row[b] = ch_cnt
        tot_ch_dec_all += ch_cnt
    ch_dec_row["Total"] = tot_ch_dec_all

    # 5. total succ Row
    tot_succ_adj_row: dict[str, Any] = {"RC/BANK NAME": "total succ"}
    tot_succ_adj_all = 0
    for b in issuers:
        adj = success_counts[b] + ch_dec_row[b]
        tot_succ_adj_row[b] = adj
        tot_succ_adj_all += adj
    tot_succ_adj_row["Total"] = tot_succ_adj_all

    # 6. total pos t Row
    tot_pos_t_row: dict[str, Any] = {"RC/BANK NAME": "total pos t"}
    tot_pos_t_all = 0
    for b in issuers:
        t_pos = tot_decline_row[b] + success_counts[b]
        tot_pos_t_row[b] = t_pos
        tot_pos_t_all += t_pos
    tot_pos_t_row["Total"] = tot_pos_t_all

    # 7. success rate Row
    rate_row: dict[str, Any] = {"RC/BANK NAME": "success rate"}
    for b in issuers:
        den = tot_pos_t_row[b]
        rate = round(tot_succ_adj_row[b] / den, 4) if den > 0 else 0.0
        rate_row[b] = rate
    rate_row["Total"] = round(tot_succ_adj_all / tot_pos_t_all, 4) if tot_pos_t_all > 0 else 0.0

    rows.append(rate_row)
    rows.append(ch_dec_row)
    rows.append(tot_succ_adj_row)
    rows.append(tot_pos_t_row)

    matrix_df = pd.DataFrame(rows)

    used_rcs = sorted(list(decline_rcs), key=lambda x: int(x) if x.isdigit() else 9999)
    desc_rows = []
    for rc in used_rcs:
        info = _LOOKUP_MAP.get(rc, {"description": "Unknown Response Code", "remark": "New or unmapped response code from network"})
        desc_rows.append({
            "Response Code": rc,
            "Description": info["description"],
            "Remark": info["remark"],
        })
    desc_df = pd.DataFrame(desc_rows)

    return matrix_df, desc_df


def build_pos_success_rate_excel(matrix_df: pd.DataFrame, desc_df: pd.DataFrame) -> bytes:
    """Build formatted Excel workbook matching POS SUCCESS RATE layout with live Excel formulas
    and tier-based conditional color coding for success rates.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "POS SUCCESS RATE"

    font_family = "Arial"

    header_title_font = Font(name=font_family, size=13, bold=True, color="FFFFFF")
    header_title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    hdr_font = Font(name=font_family, size=10, bold=True, color="1F4E78")

    data_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    thick_bottom = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )

    num_cols = len(matrix_df.columns)
    last_bank_col_idx = num_cols - 1
    last_bank_let = get_column_letter(last_bank_col_idx)
    tot_col_let = get_column_letter(num_cols)

    # Row 1: Banner Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Pos Transaction Decline Response & Success Rate Summary (Issuer View)"
    title_cell.font = header_title_font
    title_cell.fill = header_title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 3: Header
    for col_idx, col_name in enumerate(matrix_df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 22

    # Map row indices for formulas
    # Row 4 is first decline code
    rc_row_map: dict[str, int] = {}
    current_row = 4

    # Extract RC rows from matrix_df
    decline_rows_count = len(matrix_df) - 6  # 6 summary rows at the end

    for i in range(decline_rows_count):
        rc_code = str(matrix_df.iloc[i]["RC/BANK NAME"])
        rc_row_map[rc_code] = current_row
        current_row += 1

    r_start = 4
    r_end = current_row - 1 if decline_rows_count > 0 else 4

    r_tot_dec = current_row
    current_row += 1

    r_succ_pos = current_row
    current_row += 1

    r_rate = current_row
    current_row += 1

    r_ch_dec = current_row
    current_row += 1

    r_tot_succ = current_row
    current_row += 1

    r_tot_pos = current_row
    current_row += 1

    # Find row numbers for cardholder RCs that exist in the dataset
    ch_row_nums = [rc_row_map[rc] for rc in CARDHOLDER_RC if rc in rc_row_map]

    # Populate Data & Formulas into openpyxl Sheet
    for idx, r in matrix_df.iterrows():
        label = str(r["RC/BANK NAME"])

        if label in rc_row_map:
            row_idx = rc_row_map[label]
        elif label == "Total Decline":
            row_idx = r_tot_dec
        elif label == "Successful Pos T":
            row_idx = r_succ_pos
        elif label == "success rate":
            row_idx = r_rate
        elif label == "card holder rel dec":
            row_idx = r_ch_dec
        elif label == "total succ":
            row_idx = r_tot_succ
        elif label == "total pos t":
            row_idx = r_tot_pos
        else:
            row_idx = current_row

        ws.cell(row=row_idx, column=1, value=label)

        # Bank Columns (Col 2 to last_bank_col_idx)
        for col_idx in range(2, last_bank_col_idx + 1):
            col_let = get_column_letter(col_idx)
            cell = ws.cell(row=row_idx, column=col_idx)

            if label in rc_row_map:
                cell.value = int(r.iloc[col_idx - 1])
                cell.number_format = "#,##0"
                cell.font = data_font
            elif label == "Total Decline":
                cell.value = f"=SUM({col_let}{r_start}:{col_let}{r_end})" if decline_rows_count > 0 else 0
                cell.number_format = "#,##0"
                cell.font = bold_font
                cell.fill = summary_fill
            elif label == "Successful Pos T":
                cell.value = int(r.iloc[col_idx - 1])
                cell.number_format = "#,##0"
                cell.font = bold_font
                cell.fill = summary_fill
            elif label == "card holder rel dec":
                if ch_row_nums:
                    cell.value = f"={'+'.join(f'{col_let}{r_n}' for r_n in ch_row_nums)}"
                else:
                    cell.value = 0
                cell.number_format = "#,##0"
                cell.font = bold_font
                cell.fill = summary_fill
            elif label == "total succ":
                cell.value = f"={col_let}{r_succ_pos}+{col_let}{r_ch_dec}"
                cell.number_format = "#,##0"
                cell.font = bold_font
                cell.fill = summary_fill
            elif label == "total pos t":
                cell.value = f"={col_let}{r_tot_dec}+{col_let}{r_succ_pos}"
                cell.number_format = "#,##0"
                cell.font = bold_font
                cell.fill = summary_fill
            elif label == "success rate":
                cell.value = f"=IF({col_let}{r_tot_pos}>0, {col_let}{r_tot_succ}/{col_let}{r_tot_pos}, 0)"
                cell.number_format = "0.00%"
                # Apply 4-tier color styling based on pre-calculated rate
                rate_val = float(r.iloc[col_idx - 1])
                fill_style, font_style = get_rate_style(rate_val)
                cell.fill = fill_style
                cell.font = font_style

            cell.border = thick_bottom if label == "total pos t" else thin_border
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # Total Column (Col num_cols)
        tot_cell = ws.cell(row=row_idx, column=num_cols)
        if label in rc_row_map:
            tot_cell.value = f"=SUM(B{row_idx}:{last_bank_let}{row_idx})"
            tot_cell.number_format = "#,##0"
            tot_cell.font = data_font
        elif label == "Total Decline":
            tot_cell.value = f"=SUM(B{r_tot_dec}:{last_bank_let}{r_tot_dec})"
            tot_cell.number_format = "#,##0"
            tot_cell.font = bold_font
            tot_cell.fill = summary_fill
        elif label == "Successful Pos T":
            tot_cell.value = f"=SUM(B{r_succ_pos}:{last_bank_let}{r_succ_pos})"
            tot_cell.number_format = "#,##0"
            tot_cell.font = bold_font
            tot_cell.fill = summary_fill
        elif label == "card holder rel dec":
            tot_cell.value = f"=SUM(B{r_ch_dec}:{last_bank_let}{r_ch_dec})"
            tot_cell.number_format = "#,##0"
            tot_cell.font = bold_font
            tot_cell.fill = summary_fill
        elif label == "total succ":
            tot_cell.value = f"=SUM(B{r_tot_succ}:{last_bank_let}{r_tot_succ})"
            tot_cell.number_format = "#,##0"
            tot_cell.font = bold_font
            tot_cell.fill = summary_fill
        elif label == "total pos t":
            tot_cell.value = f"=SUM(B{r_tot_pos}:{last_bank_let}{r_tot_pos})"
            tot_cell.number_format = "#,##0"
            tot_cell.font = bold_font
            tot_cell.fill = summary_fill
        elif label == "success rate":
            # Total Column success rate formula `=Total_total_succ / Total_total_pos_t`
            tot_cell.value = f"=IF({tot_col_let}{r_tot_pos}>0, {tot_col_let}{r_tot_succ}/{tot_col_let}{r_tot_pos}, 0)"
            tot_cell.number_format = "0.00%"
            tot_rate_val = float(r.iloc[num_cols - 1])
            t_fill_style, t_font_style = get_rate_style(tot_rate_val)
            tot_cell.fill = t_fill_style
            tot_cell.font = t_font_style

        tot_cell.border = thick_bottom if label == "total pos t" else thin_border
        tot_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Column A Header cell formatting
        hdr_c = ws.cell(row=row_idx, column=1)
        is_summary_label = label in ("Total Decline", "Successful Pos T", "card holder rel dec", "total succ", "total pos t")
        hdr_c.font = bold_font if (is_summary_label or label == "success rate") else data_font
        if label == "success rate":
            tot_rate_val = float(r.iloc[num_cols - 1])
            f_st, _ = get_rate_style(tot_rate_val)
            hdr_c.fill = f_st
        elif is_summary_label:
            hdr_c.fill = summary_fill
        hdr_c.border = thick_bottom if label == "total pos t" else thin_border

        ws.row_dimensions[row_idx].height = 19

    # Add openpyxl dynamic conditional formatting rules to row r_rate
    rate_range_str = f"B{r_rate}:{tot_col_let}{r_rate}"
    rule_green = CellIsRule(operator="greaterThanOrEqual", formula=["0.97"], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), font=Font(name=font_family, size=10, bold=True, color="006100"))
    rule_yellow = CellIsRule(operator="between", formula=["0.86", "0.969999"], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), font=Font(name=font_family, size=10, bold=True, color="9C6500"))
    rule_lyellow = CellIsRule(operator="between", formula=["0.79", "0.859999"], fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), font=Font(name=font_family, size=10, bold=True, color="7F6000"))
    rule_red = CellIsRule(operator="lessThanOrEqual", formula=["0.789999"], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), font=Font(name=font_family, size=10, bold=True, color="9C0006"))

    ws.conditional_formatting.add(rate_range_str, rule_green)
    ws.conditional_formatting.add(rate_range_str, rule_yellow)
    ws.conditional_formatting.add(rate_range_str, rule_lyellow)
    ws.conditional_formatting.add(rate_range_str, rule_red)

    # Row + 3: Response Code Reference Table Header
    ref_start_row = r_tot_pos + 3
    ws.cell(row=ref_start_row, column=1, value="Response Code Description & Remarks Reference Table").font = Font(name=font_family, size=11, bold=True)
    ref_start_row += 1

    desc_cols = ["Response Code", "Description", "Remark"]
    for c_idx, c_name in enumerate(desc_cols, start=1):
        cell = ws.cell(row=ref_start_row, column=c_idx, value=c_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin_border
    ws.row_dimensions[ref_start_row].height = 20
    ref_start_row += 1

    for _, d_row in desc_df.iterrows():
        ws.cell(row=ref_start_row, column=1, value=str(d_row["Response Code"])).alignment = Alignment(horizontal="center")
        ws.cell(row=ref_start_row, column=2, value=str(d_row["Description"])).alignment = Alignment(horizontal="left")
        ws.cell(row=ref_start_row, column=3, value=str(d_row["Remark"])).alignment = Alignment(horizontal="left")

        for c_idx in range(1, 4):
            c = ws.cell(row=ref_start_row, column=c_idx)
            c.font = data_font
            c.border = thin_border

        ws.row_dimensions[ref_start_row].height = 18
        ref_start_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
