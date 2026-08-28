import io

import pytest
from openpyxl import Workbook

import app as app_module
from merger import CANONICAL_COLUMNS, POS_CANONICAL_COLUMNS


def _make_xlsx(filename: str = "export.xlsx") -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Worksheet"
    ws.append(
        ["ACQUIRER", "ISSUER", "PAN", "TRANS_DATE", "TIME", "TRANS_TYPE",
         "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM", "ADDRESS_NAME"]
    )
    ws.append(["Commercial Bank", "Abyssinia Bank", "4006780*****5775", 20260813,
               "01:20:54", "Purchase", 1216, 801, 0, 260813000058603840,
               "000001001632", "NURHUSSEN YASSIN OMER"])
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue()


@pytest.fixture
def client():
    app_module.app.config["TESTING"] = True
    return app_module.app.test_client()


def test_index_page(client):
    resp = client.get("/")
    assert resp.status_code == 200
    assert b"POS &amp; ATM Report Merger" in resp.data
    # the mode picker offers all three report types
    assert b'data-mode="pos_decline"' in resp.data
    assert b'data-mode="pos"' in resp.data
    assert b'data-mode="atm"' in resp.data
    # breadcrumbs are present so the user can go back to the picker
    assert b'id="breadcrumbs"' in resp.data
    assert b'id="crumb-home"' in resp.data


def test_merge_and_download_roundtrip(client):
    filename, data = _make_xlsx()
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(data), filename)},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["total_rows"] == 1
    assert payload["columns"] == CANONICAL_COLUMNS
    assert payload["filename"].endswith("Merged.xlsx")
    token = payload["token"]

    dl = client.get(f"/download/{token}")
    assert dl.status_code == 200
    assert dl.headers["Content-Disposition"].startswith("attachment")
    assert dl.data[:2] == b"PK"  # valid xlsx zip

    # token persists for filter-download (cache is kept until TTL expiry)


def test_merge_multiple_files(client):
    f1, d1 = _make_xlsx("a.xlsx")
    f2, d2 = _make_xlsx("b.xlsx")
    resp = client.post(
        "/merge",
        data={
            "files": [
                (io.BytesIO(d1), f1),
                (io.BytesIO(d2), f2),
            ]
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["total_rows"] == 2


def test_merge_reports_column_mismatch_warnings(client):
    filename, data = _make_xlsx("full.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM"])
    ws.append(["Bank A", "Bank B", "1111*****2222", 20260813, "10:00:00",
               "Purchase", 100, 901, 0, 123456, "REF-1"])
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/merge",
        data={
            "files": [
                (io.BytesIO(data), filename),
                (io.BytesIO(buf.getvalue()), "no_merchant.xlsx"),
            ]
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["total_rows"] == 2
    assert any("no_merchant.xlsx" in w and "MERCHANT" in w for w in payload["warnings"])


def _make_atm_xlsx(filename: str = "atm.xlsx") -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN", "UTRNNO",
               "TERMINAL_ID", "ADDRESS_NAME"])
    ws.append(["", "Commercial Bank", "4583006*****1077", 20260815, "0:0:20",
               "ATM cash withdrawal", 2000, 230, "-1", "002724804961", 2724804961,
               "AE005001", "Siket Bank"])
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue()


ATM_COLUMNS = ["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN", "UTRNNO",
               "TERMINAL_ID", "ADDRESS_NAME"]


POS_DAILY_COLUMNS = ["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
                     "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN",
                     "TERMINAL_ID", "ADDRESS"]


def _make_pos_daily_xlsx(filename: str = "pos.xlsx") -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "UTRNNO", "RRN",
               "TERMINAL_ID", "ADDRESS_NAME"])
    ws.append(["Dashen Bank", "Commercial Bank", "4583006*****7109", 20260815,
               "00:24:38", "Purchase", 305, 230, "-1", 260815000066964000,
               "622721673223", "TYLUSC02", "YUMMY LUSCIOUS CAFE PLC"])
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue()


def test_merge_pos_mode(client):
    filename, data = _make_pos_daily_xlsx()
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(data), filename), "mode": "pos"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["mode"] == "pos"
    # output keeps only the sample columns - the extra UTRNNO column removed
    assert payload["columns"] == POS_DAILY_COLUMNS
    assert "UTRNNO" not in payload["columns"]
    assert payload["total_rows"] == 1
    assert payload["filename"].startswith("Daily_Tranaction_Report_SmartVista_POS")
    # the removed column is reported in the per-file diagnostics
    info = payload["per_file"][0]
    assert info["extra_columns"] == ["UTRNNO"]
    dl = client.get(f"/download/{payload['token']}")
    assert dl.status_code == 200
    assert dl.data[:2] == b"PK"


def test_merge_atm_mode(client):
    filename, data = _make_atm_xlsx()
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(data), filename), "mode": "atm"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["mode"] == "atm"
    assert payload["columns"] == ATM_COLUMNS
    assert payload["total_rows"] == 1
    assert payload["filename"].startswith("Daily_Tranaction_Report_SmartVista_ATM")
    dl = client.get(f"/download/{payload['token']}")
    assert dl.status_code == 200
    assert dl.data[:2] == b"PK"


def test_merge_rejects_unknown_mode(client):
    filename, data = _make_xlsx()
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(data), filename), "mode": "bogus"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400


def test_merge_rejects_non_excel(client):
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(b"not excel at all"), "bad.xlsx")},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 400
    assert "error" in resp.get_json()


def test_merge_rejects_empty_upload(client):
    resp = client.post("/merge", data={}, content_type="multipart/form-data")
    assert resp.status_code == 400


POS_SUCCESS_COLUMNS = ["ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME",
                       "TRANS_TYPE", "AMOUNT", "RESP_CODE", "REFNUM",
                       "UTRNNO", "MERCHANT"]


def _make_pos_success_xlsx(filename: str = "pos_success.xlsx") -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Worksheet"
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRANS_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REFNUM", "FE_UTRNNO", "ADDRESS_NAME"])
    ws.append(["Wegagen Bank", "Commercial Bank", "4585716*****2369", 20260813,
               "01:13:11", "Purchase", 2040, -1, "622422564685",
               "260813000058601088", "Yechaka Bunna PLC"])
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue()


def test_merge_pos_success_mode(client):
    filename, data = _make_pos_success_xlsx()
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(data), filename), "mode": "pos_success"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["mode"] == "pos_success"
    assert payload["columns"] == POS_SUCCESS_COLUMNS
    assert payload["total_rows"] == 1
    assert payload["filename"].startswith("POS_Transaction_SVFE_Report")
    dl = client.get(f"/download/{payload['token']}")
    assert dl.status_code == 200
    assert dl.data[:2] == b"PK"


def test_merge_pos_success_multiple_files(client):
    f1, d1 = _make_pos_success_xlsx("a.xlsx")
    f2, d2 = _make_pos_success_xlsx("b.xlsx")
    resp = client.post(
        "/merge",
        data={
            "files": [
                (io.BytesIO(d1), f1),
                (io.BytesIO(d2), f2),
            ],
            "mode": "pos_success",
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.get_json()["total_rows"] == 2


QR_COLUMNS = ["DESTINATION_BANK", "SOURCE_BANK", "TRX_DATE", "DBTR_ACCT",
              "CDTR_ACCT", "AMOUNT", "TX_ID", "STATUS"]


def _make_qr_xlsx(filename: str = "qr.xlsx") -> tuple[str, bytes]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["DESTINATION_BANK", "SOURCE_BANK", "TRX_DATE", "DBTR_ACCT",
               "CDTR_ACCT", "AMOUNT", "TX_ID", "STATUS"])
    ws.append(["Awash Bank", "Cooperative Bank of Oromia", 20250707,
               "251915180606", "014251099975200", 77600,
               "CBORETAA1455091289", "PROCESSED"])
    buf = io.BytesIO()
    wb.save(buf)
    return filename, buf.getvalue()


def test_index_page_offers_qr_mode(client):
    resp = client.get("/")
    assert resp.status_code == 200
    assert b'data-mode="qr"' in resp.data
    assert b">QR</span>" in resp.data


def test_merge_qr_mode(client):
    filename, data = _make_qr_xlsx()
    resp = client.post(
        "/merge",
        data={"files": (io.BytesIO(data), filename), "mode": "qr"},
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["mode"] == "qr"
    assert payload["columns"] == QR_COLUMNS
    assert payload["total_rows"] == 1
    assert payload["filename"].startswith("QR_Export")
    assert payload["resp_counts"] == {"PROCESSED": 1}
    dl = client.get(f"/download/{payload['token']}")
    assert dl.status_code == 200
    assert dl.data[:2] == b"PK"


def test_merge_qr_mode_sort(client):
    filename, data = _make_qr_xlsx()
    resp = client.post(
        "/merge",
        data={
            "files": (io.BytesIO(data), filename),
            "mode": "qr",
            "sort_by": "AMOUNT",
            "sort_dir": "desc",
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["sort_by"] == "AMOUNT"
    assert payload["sort_dir"] == "desc"
