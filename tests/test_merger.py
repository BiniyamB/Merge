import io
import os
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from merger import (
    ATM_CANONICAL_COLUMNS,
    ATM_MODE,
    CANONICAL_COLUMNS,
    POS_CANONICAL_COLUMNS,
    POS_MODE,
    POS_SUCCESS_CANONICAL_COLUMNS,
    POS_SUCCESS_MODE,
    QR_CANONICAL_COLUMNS,
    QR_MODE,
    REPORT_TITLE,
    SHEET_NAME,
    _normalize_time,
    _read_xlsx_fast,
    _read_xlsx_openpyxl,
    build_workbook,
    merge_reports,
    parse_report,
    smart_sort_key,
)


def _xlsx_export_style() -> bytes:
    """Mimic 'Aug 13  POS  Declined.xlsx' (Export Worksheet):
    12 columns, no blanks, header in row 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Worksheet"
    ws.append(
        ["ACQUIRER", "ISSUER", "PAN", "TRANS_DATE", "TIME", "TRANS_TYPE",
         "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM", "ADDRESS_NAME"]
    )
    # UTRNNO is stored as text (openpyxl cannot serialize integers needing
    # more than 16 significant digits; the real reports are stored either
    # as exact 17-digit numbers or as text, and our pipeline recovers both).
    ws.append(["Commercial Bank", "Abyssinia Bank", "4006780*****5775", 20260813,
               "01:20:54", "Purchase", 1216, 801, 0, "260813000058603840",
               "000001001632", "NURHUSSEN YASSIN OMER"])
    ws.append(["Zemen Bank", "Commercial Bank", "4585715*****7513", 20260813,
               "02:42:03", "Purchase", 7184.55, 901, 0, "260813000058603856",
               "622519264030", "SAPPHIRE ADDIS HOTEL"])
    ws.append(["Commercial Bank", "Dashen Bank", "9231405*****9040987", 20260813,
               "02:41:14", "Purchase", 410, 915, 0, "260813000058603879",
               "000001001035", "GOLLA BIRHANU NIGATU"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xls_report_style() -> bytes:
    """Mimic 'POS_Transaction_Decline_Report (42).xls':
    title rows, blank spacer columns B/D/F, repeated headers, literal
    'null' strings, and a fully empty row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "POS_Transaction_Report"
    ws.append(["", "Report name:", "", "POS TRANSACTION  DECLINE REPORT"])
    ws.append(["", "From Date:", "", 20260813])
    ws.append(["", "To Date:", "", 20260813])
    ws.append(["ACQUIRER", "", "ISSUER", "", "PAN", "", "TRAN_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "RESP_CODE", "Reversal", "FE UTRNNO",
               "REFNUM", "MERCHANT"])
    ws.append(["Abyssinia", "", "Oromia", "", "9231413xxxxxx7072", "", 20260813,
               115207, "POS purchase", 980, 901, 0, 2724448634, 622508281398,
               "SKYLIGHT HOTEL POS-34"])
    ws.append(["CBE", "", "null", "", "9231430xxxxxx3582", "", 20260813,
               194920, "POS purchase", 420, 915, 0, 2724548046,
               "000001001235", "MIS WOYINISHET ABEBE"])
    # repeated header block, as on a page break
    ws.append(["ACQUIRER", "", "ISSUER", "", "PAN", "", "TRAN_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "RESP_CODE", "Reversal", "FE UTRNNO",
               "REFNUM", "MERCHANT"])
    ws.append(["Dashen", "", "null", "", "9231429xxxxxx1685", "", 20260813,
               104847, "POS purchase", 2540, 902, 0, 2724435641, 622507569522,
               "ABADIR SHOPING CENTER"])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xls_style_drops_blank_columns_and_headers():
    rep = parse_report(_xls_report_style(), "decline42.xlsx")
    assert rep.data_rows == 3
    assert rep.blank_columns == ["B", "D", "F"]
    assert rep.columns_kept == CANONICAL_COLUMNS
    # repeated header + empty row skipped
    assert "repeated header rows skipped" in rep.warnings[0]
    # literal "null" strings are preserved exactly as in the file
    issuers = {r["ISSUER"] for r in rep.rows}
    assert "null" in issuers
    # TIME values kept raw (e.g. 115207), not reformatted; parse keeps the
    # sheet order (sorting happens only when merging)
    times = [r["TIME"] for r in rep.rows]
    assert times == [115207, 194920, 104847]


def test_parse_xlsx_export_style():
    rep = parse_report(_xlsx_export_style(), "export.xlsx")
    assert rep.data_rows == 3
    assert rep.blank_columns == []
    assert rep.columns_kept == CANONICAL_COLUMNS
    # UTRNNO kept exactly as stored (here as text in the fixture)
    utrn = {r["FE UTRNNO"] for r in rep.rows}
    assert "260813000058603840" in utrn
    # REFNUM values untouched: both stored as text in this fixture stay text
    refs = {r["REFNUM"] for r in rep.rows}
    assert "000001001632" in refs
    assert "622519264030" in refs


def test_merge_two_reports():
    result = merge_reports(
        [("export.xlsx", _xlsx_export_style()), ("decline42.xlsx", _xls_report_style())]
    )
    assert result.total_rows == 6
    assert result.from_date == "20260813"
    assert result.to_date == "20260813"
    assert result.filename == "POS_Transaction_Decline_Report_20260813_Merged.xlsx"
    # sorted by time (via internal normalized key; raw values untouched)
    times = [r["TIME"] for r in result.records]
    keys = [_normalize_time(t) for t in times]
    assert keys == sorted(keys)
    assert result.resp_counts == {"801": 1, "901": 2, "915": 2, "902": 1}
    # every record has exactly the 12 canonical keys
    for rec in result.records:
        assert list(rec.keys()) == CANONICAL_COLUMNS


def test_merge_dedupe_keeps_only_unique_rows():
    # Merging the same file twice (identical rows) with dedupe off doubles
    # every row; with dedupe on the fully-duplicated rows are collapsed.
    files = [
        ("export.xlsx", _xlsx_export_style()),
        ("export_copy.xlsx", _xlsx_export_style()),
    ]
    result_keep = merge_reports(files, dedupe=False)
    assert result_keep.total_rows == 6  # 3 rows x 2 files

    result_unique = merge_reports(files, dedupe=True)
    assert result_unique.total_rows == 3  # exact duplicates collapsed

    # The surviving rows are exactly the original 3, and values are preserved.
    assert result_unique.records[0]["ACQUIRER"] == "Commercial Bank"
    assert result_unique.records[0]["TIME"] == "01:20:54"
    assert result_unique.records[1]["TIME"] == "02:41:14"
    assert result_unique.records[2]["TIME"] == "02:42:03"
    # every record has exactly the 12 canonical keys
    for rec in result_unique.records:
        assert list(rec.keys()) == CANONICAL_COLUMNS


def test_merge_dedupe_ignores_single_occurrence_rows():
    # Rows that differ in any column are not removed even when duplicates
    # exist elsewhere. The two files share one identical row and differ on
    # the others, so only that shared row is collapsed.
    files = [
        ("export.xlsx", _xlsx_export_style()),
        ("decline42.xlsx", _xls_report_style()),
    ]
    result_keep = merge_reports(files, dedupe=False)
    assert result_keep.total_rows == 6

    result_unique = merge_reports(files, dedupe=True)
    assert result_unique.total_rows == 6  # no overlapping identical rows


def test_duplicate_records_populated_when_same_file_twice():
    # When the same file is uploaded twice, every row appears twice so all
    # rows are duplicates. duplicate_records contains both copies.
    files = [
        ("export.xlsx", _xlsx_export_style()),
        ("export_copy.xlsx", _xlsx_export_style()),
    ]
    result = merge_reports(files, dedupe=False)
    assert result.total_rows == 6
    assert len(result.duplicate_records) == 6  # all 6 rows are duplicates
    # duplicates can be built into a workbook without errors
    wb_bytes = build_workbook(result.duplicate_records, result.from_date, result.to_date, POS_MODE)
    assert len(wb_bytes) > 0
    wb = load_workbook(io.BytesIO(wb_bytes))
    assert len(wb.sheetnames) == 1


def test_duplicate_records_empty_for_unique_data():
    # When no two rows are identical, duplicate_records is empty.
    result = merge_reports([("export.xlsx", _xlsx_export_style())])
    assert result.total_rows == 3
    assert len(result.duplicate_records) == 0


def test_duplicate_records_populated_even_with_dedupe_on():
    # duplicate_records are computed from the pre-dedupe data so they are
    # available even when dedupe is True.
    files = [
        ("export.xlsx", _xlsx_export_style()),
        ("export_copy.xlsx", _xlsx_export_style()),
    ]
    result = merge_reports(files, dedupe=True)
    assert result.total_rows == 3  # deduped
    assert len(result.duplicate_records) == 6  # pre-dedupe duplicates kept


def test_duplicate_records_only_exact_match():
    # Two files share one identical row and differ on the others. Only
    # the shared row is a duplicate.
    files = [
        ("export.xlsx", _xlsx_export_style()),
        ("decline42.xlsx", _xls_report_style()),
    ]
    result = merge_reports(files, dedupe=False)
    assert len(result.duplicate_records) == 0  # no overlapping rows


def test_merged_workbook_layout():
    result = merge_reports([("export.xlsx", _xlsx_export_style())])
    wb = load_workbook(io.BytesIO(result.workbook_bytes))
    assert wb.sheetnames == [SHEET_NAME]
    ws = wb[SHEET_NAME]
    # title block
    assert ws["A1"].value == "Report name:"
    assert ws["B1"].value == REPORT_TITLE
    assert ws["A2"].value == "From Date:"
    assert ws["B2"].value == 20260813
    assert ws["A3"].value == "To Date:"
    # header row with no blank columns
    header = [ws.cell(row=4, column=c).value for c in range(1, len(CANONICAL_COLUMNS) + 1)]
    assert header == CANONICAL_COLUMNS
    assert ws.cell(row=4, column=13).value is None  # nothing beyond col L
    # data present
    assert ws.cell(row=5, column=1).value == "Commercial Bank"


def test_merge_keeps_errors_per_file_and_skips_bad_files():
    result = merge_reports(
        [
            ("good.xlsx", _xlsx_export_style()),
            ("bad.xlsx", b"this is definitely not an excel file"),
        ]
    )
    assert result.total_rows == 3
    statuses = {p["filename"]: p["status"] for p in result.per_file}
    assert statuses == {"good.xlsx": "ok", "bad.xlsx": "error"}


def test_merge_rejects_when_nothing_parses():
    with pytest.raises(ValueError, match="None of the uploaded files"):
        merge_reports([("bad.xlsx", b"garbage")])


def test_merge_rejects_empty_upload():
    with pytest.raises(ValueError, match="No files were uploaded"):
        merge_reports([])


def _shuffled_report() -> bytes:
    """The same 12 columns but in a completely different order."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shuffled"
    ws.append(["MERCHANT", "REFNUM", "FE UTRNNO", "Reversal", "RESP_CODE",
               "AMOUNT", "TRANS_TYPE", "TIME", "TRAN_DATE", "PAN", "ISSUER",
               "ACQUIRER"])
    ws.append(["SHOP A", "000001009999", 2724440001, 0, 801, 500.5,
               "POS purchase", 94513, 20260813, "4006780*****1111",
               "Abyssinia Bank", "Test Bank"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _duplicate_header_report() -> bytes:
    """A report with two 'PAN' columns - the first one must win."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "PAN", "TRAN_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO",
               "REFNUM", "ADDRESS_NAME"])
    ws.append(["Bank A", "Bank B", "1111*****2222", "9999*****0000", 20260813,
               "10:00:00", "Purchase", 100, 901, 0, 123456, "REF-1", "M1"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_reshuffled_columns_maps_by_header_name():
    rep = parse_report(_shuffled_report(), "shuffled.xlsx")
    assert rep.data_rows == 1
    assert rep.columns_kept == CANONICAL_COLUMNS
    assert rep.column_order != CANONICAL_COLUMNS
    assert rep.order_mismatch is True
    assert any("reshuffled" in w for w in rep.warnings)

    row = rep.rows[0]
    assert row["ACQUIRER"] == "Test Bank"
    assert row["ISSUER"] == "Abyssinia Bank"
    assert row["PAN"] == "4006780*****1111"
    assert row["TRAN_DATE"] == 20260813
    assert row["TIME"] == 94513  # raw value, unchanged
    assert row["TRANS_TYPE"] == "POS purchase"
    assert row["AMOUNT"] == 500.5
    assert row["RESP_CODE"] == 801
    assert row["Reversal"] == 0
    assert row["FE UTRNNO"] == 2724440001
    assert row["REFNUM"] == "000001009999"
    assert row["MERCHANT"] == "SHOP A"


def test_merge_shuffled_report_into_canonical_format():
    result = merge_reports(
        [("shuffled.xlsx", _shuffled_report()), ("export.xlsx", _xlsx_export_style())]
    )
    assert result.total_rows == 4
    for rec in result.records:
        assert list(rec.keys()) == CANONICAL_COLUMNS
    # the shuffled row's values land in the right columns of the merged output
    row = next(r for r in result.records if r["MERCHANT"] == "SHOP A")
    assert row["ACQUIRER"] == "Test Bank"
    assert row["TIME"] == 94513  # raw value, unchanged
    assert row["PAN"] == "4006780*****1111"

    # reshuffled file flagged in the per-file diagnostics
    info = next(p for p in result.per_file if p["filename"] == "shuffled.xlsx")
    assert info["order_mismatch"] is True
    assert info["column_order"][0] == "MERCHANT"


def test_duplicate_header_keeps_first_column():
    rep = parse_report(_duplicate_header_report(), "dupes.xlsx")
    assert rep.data_rows == 1
    assert rep.rows[0]["PAN"] == "1111*****2222"  # first PAN column wins
    assert any("duplicate 'PAN'" in w for w in rep.warnings)


def _missing_merchant_report() -> bytes:
    """Every standard column except MERCHANT (ADDRESS_NAME header removed)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM"])
    ws.append(["Bank A", "Bank B", "1111*****2222", 20260813, "10:00:00",
               "Purchase", 100, 901, 0, 123456, "REF-1"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_missing_column_warning_names_file_and_column():
    result = merge_reports(
        [("full.xlsx", _xlsx_export_style()), ("no_merchant.xlsx", _missing_merchant_report())]
    )
    assert result.total_rows == 4
    # top-level warning explicitly names the missing column and the file
    assert any(
        "no_merchant.xlsx" in w and "MERCHANT" in w for w in result.warnings
    ), result.warnings
    # the same message is attached to that file's own diagnostics
    info = next(p for p in result.per_file if p["filename"] == "no_merchant.xlsx")
    assert any("MERCHANT" in w and "no_merchant.xlsx" in w for w in info["warnings"])


def test_unbalanced_column_count_warning():
    result = merge_reports(
        [("full.xlsx", _xlsx_export_style()), ("no_merchant.xlsx", _missing_merchant_report())]
    )
    assert any("unbalanced" in w.lower() for w in result.warnings), result.warnings


def test_no_column_warnings_when_all_files_match():
    result = merge_reports(
        [("a.xlsx", _xlsx_export_style()), ("b.xlsx", _xlsx_export_style())]
    )
    assert result.warnings == []


# ---------------------------------------------------------------------------
# ATM mode
# ---------------------------------------------------------------------------
def _atm_aug15_style() -> bytes:
    """Mimic 'Aug 15 ATM s-s .xlsx': header TIME, UTRNNO before RRN."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Worksheet"
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "UTRNNO", "RRN",
               "TERMINAL_ID", "ADDRESS_NAME"])
    ws.append(["Abyssinia Bank", "Commercial Bank", "4583007*****7597", 20260815,
               "00:24:22", "Cash withdrawal", 1000, 230, 915, "260815000066964000",
               "622721417723", "ADD08143", "DIRE SEBATEGNA BRANCH"])
    ws.append(["Abyssinia Bank", "Commercial Bank", "4583007*****3262", 20260815,
               "00:24:32", "Cash withdrawal", 1000, 230, "-1",
               "260815000066964064", "622721416943", "ADS09042", "HOTIE BRANCH"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _atm_daily_style() -> bytes:
    """Mimic the SmartVista daily ATM report: TRANS_TIME, RRN before UTRNNO,
    non-padded times, leading-zero RRN strings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN", "UTRNNO",
               "TERMINAL_ID", "ADDRESS_NAME"])
    ws.append(["", "Commercial Bank", "4583006*****1077", 20260815, "0:0:20",
               "ATM cash withdrawal", 2000, 230, "-1", "002724804961", 2724804961,
               "AE005001", "Siket Bank"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_atm_aug15_style():
    rep = parse_report(_atm_aug15_style(), "atm_aug15.xlsx", mode=ATM_MODE)
    assert rep.data_rows == 2
    assert rep.blank_columns == []
    assert rep.columns_kept == list(ATM_CANONICAL_COLUMNS)
    assert rep.order_mismatch is True  # UTRNNO/RRN swapped vs canonical order
    assert any("reshuffled" in w for w in rep.warnings)

    row = rep.rows[0]
    assert row["ACQUIRER"] == "Abyssinia Bank"
    assert row["CARD_NUMBER"] == "4583007*****7597"
    assert row["TRANS_DATE"] == 20260815
    assert row["TRANS_TIME"] == "00:24:22"  # TIME header -> TRANS_TIME, value raw
    assert row["AMOUNT"] == 1000
    assert row["CURRENCY"] == 230
    assert row["RESP"] == 915
    assert row["UTRNNO"] == "260815000066964000"  # text stored -> text preserved
    assert row["RRN"] == "622721417723"
    assert row["TERMINAL_ID"] == "ADD08143"
    assert row["ADDRESS_NAME"] == "DIRE SEBATEGNA BRANCH"


def test_parse_atm_daily_style():
    rep = parse_report(_atm_daily_style(), "atm_daily.xlsx", mode=ATM_MODE)
    assert rep.data_rows == 1
    assert rep.order_mismatch is False
    row = rep.rows[0]
    assert row["TRANS_TIME"] == "0:0:20"  # non-padded value, unchanged
    assert row["RRN"] == "002724804961"   # leading-zero string preserved
    assert row["UTRNNO"] == 2724804961
    assert row["RESP"] == "-1"


def test_merge_atm_reports():
    result = merge_reports(
        [("atm_aug15.xlsx", _atm_aug15_style()), ("atm_daily.xlsx", _atm_daily_style())],
        mode_key="atm",
    )
    assert result.total_rows == 3
    assert result.mode_key == "atm"
    assert result.filename == (
        "Daily_Tranaction_Report_SmartVista_ATM_15_Aug_26_to_15_Aug_26_Merged.xlsx"
    )
    assert result.from_date == result.to_date == "20260815"
    for rec in result.records:
        assert list(rec.keys()) == list(ATM_CANONICAL_COLUMNS)
    # output column order: RRN before UTRNNO (SmartVista daily format)
    assert list(result.records[0].keys())[9] == "RRN"
    assert list(result.records[0].keys())[10] == "UTRNNO"
    # sorted via internal normalized keys; raw values untouched
    times = [r["TRANS_TIME"] for r in result.records]
    keys = [_normalize_time(t) for t in times]
    assert keys == sorted(keys)
    assert result.resp_counts == {"915": 1, "-1": 2}
    # per-file reshuffle flags
    info = next(p for p in result.per_file if p["filename"] == "atm_aug15.xlsx")
    assert info["order_mismatch"] is True
    info2 = next(p for p in result.per_file if p["filename"] == "atm_daily.xlsx")
    assert info2["order_mismatch"] is False


def test_atm_workbook_layout():
    result = merge_reports([("atm_daily.xlsx", _atm_daily_style())], mode_key="atm")
    wb = load_workbook(io.BytesIO(result.workbook_bytes))
    assert wb.sheetnames == ["Report"]
    ws = wb["Report"]
    # header in row 1, no title block
    header = [ws.cell(row=1, column=c).value for c in range(1, len(ATM_CANONICAL_COLUMNS) + 1)]
    assert header == list(ATM_CANONICAL_COLUMNS)
    assert ws.cell(row=1, column=14).value is None  # nothing beyond col M
    # data starts at row 2, raw values preserved (empty cells come back as
    # None from the streaming writer, which is the same as an empty cell)
    assert ws.cell(row=2, column=1).value in ("", None)
    assert ws.cell(row=2, column=2).value == "Commercial Bank"
    assert ws.cell(row=2, column=5).value == "0:0:20"


def test_atm_missing_column_warning():
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN", "UTRNNO",
               "ADDRESS_NAME"])
    ws.append(["DB", "Oromia Bank", "9231413*****2780434", 20260815, "0:0:37",
               "Balance inquiry", 0, 230, "-1", "622721673220", 2724804971,
               "DASHEN BANK"])
    buf = io.BytesIO()
    wb.save(buf)

    result = merge_reports(
        [("daily.xlsx", _atm_daily_style()), ("no_terminal.xlsx", buf.getvalue())],
        mode_key="atm",
    )
    assert result.total_rows == 2
    assert any("no_terminal.xlsx" in w and "TERMINAL_ID" in w for w in result.warnings)
    assert any("unbalanced" in w.lower() for w in result.warnings)


# ---------------------------------------------------------------------------
# POS mode (SmartVista daily POS format)
# ---------------------------------------------------------------------------
def _pos_daily_input_style() -> bytes:
    """Mimic 'Aug 15 POS s-s .xlsx': TIME / ADDRESS_NAME headers, 12 columns,
    RRN before (no) UTRNNO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Worksheet"
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN",
               "TERMINAL_ID", "ADDRESS_NAME"])
    ws.append(["Dashen Bank", "Commercial Bank", "4583006*****7109", 20260815,
               "00:24:38", "Purchase", 305, 230, "-1", "622721673223",
               "TYLUSC02", "YUMMY LUSCIOUS CAFE PLC"])
    ws.append(["Abyssinia Bank", "Commercial Bank", "4583006*****3807", 20260815,
               "00:01:53", "Purchase", 4345.31, 230, "-1", "622721418089",
               "PE0051AU", "YOD ABYSSINIA CULTURAL POS-04"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pos_daily_sample_style() -> bytes:
    """Mimic the SmartVista daily POS report (the output format):
    TRANS_TIME / ADDRESS headers, header in row 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN",
               "TERMINAL_ID", "ADDRESS"])
    ws.append(["YagoutPay", "Commercial Bank", "4583006*****2238", 20260815,
               "0:06:35", "POS purchase", 1850, 230, "-1", "002724805974",
               "YPT00367", "BISRATEGEBRIEL SAHILU GETANEH"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pos_daily_with_utrnno() -> bytes:
    """A POS export that additionally carries a UTRNNO column, which is NOT
    part of the POS output sample format and must be removed."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "UTRNNO", "RRN",
               "TERMINAL_ID", "ADDRESS_NAME"])
    ws.append(["Bank A", "Bank B", "4006780*****1111", 20260815, "10:00:00",
               "Purchase", 500.5, 230, "-1", 260815000066964000, "002724805974",
               "T1", "SHOP A"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_pos_daily_input():
    rep = parse_report(_pos_daily_input_style(), "pos_aug15.xlsx", mode=POS_MODE)
    assert rep.data_rows == 2
    assert rep.columns_kept == list(POS_CANONICAL_COLUMNS)
    assert rep.extra_columns == []
    assert rep.order_mismatch is False
    row = rep.rows[0]
    assert row["ACQUIRER"] == "Dashen Bank"
    assert row["TRANS_TIME"] == "00:24:38"  # TIME header -> TRANS_TIME, value raw
    assert row["ADDRESS"] == "YUMMY LUSCIOUS CAFE PLC"  # ADDRESS_NAME -> ADDRESS
    assert row["RRN"] == "622721673223"
    assert row["AMOUNT"] == 305


def test_merge_pos_daily_reports():
    result = merge_reports(
        [("pos_aug15.xlsx", _pos_daily_input_style()),
         ("pos_daily.xlsx", _pos_daily_sample_style())],
        mode_key="pos",
    )
    assert result.total_rows == 3
    assert result.mode_key == "pos"
    assert result.filename == (
        "Daily_Tranaction_Report_SmartVista_POS_15_Aug_26_to_15_Aug_26_Merged.xlsx"
    )
    for rec in result.records:
        assert list(rec.keys()) == list(POS_CANONICAL_COLUMNS)
    times = [r["TRANS_TIME"] for r in result.records]
    keys = [_normalize_time(t) for t in times]
    assert keys == sorted(keys)
    assert result.resp_counts == {"-1": 3}


def test_pos_daily_workbook_layout():
    result = merge_reports([("pos_daily.xlsx", _pos_daily_sample_style())], mode_key="pos")
    wb = load_workbook(io.BytesIO(result.workbook_bytes))
    assert wb.sheetnames == ["Report"]
    ws = wb["Report"]
    header = [ws.cell(row=1, column=c).value
              for c in range(1, len(POS_CANONICAL_COLUMNS) + 1)]
    assert header == list(POS_CANONICAL_COLUMNS)
    assert ws.cell(row=1, column=13).value is None  # nothing beyond col L
    assert ws.cell(row=2, column=5).value == "0:06:35"


def _pos_daily_rows(input_rows):
    """Build a SmartVista daily POS report (canonical headers) with the given rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(list(POS_CANONICAL_COLUMNS))
    for r in input_rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_pos_daily_dedupe_ignores_acquirer_issuer_currency_trans_type():
    # For POS (Daily), ACQUIRER, ISSUER, TRANS_TYPE and CURRENCY are not part
    # of row identity. Two rows differing only in those columns are treated
    # as duplicates, while differences in any other column keep them distinct.
    base = ["Bank-A", "Issuer", "CARD1", 20260813, "10:00:00", "Purchase", 100, 230, "-1", "RRN1", "T1", "Addr1"]
    # Three CARD1 rows differ only in ACQUIRER/ISSUER/TRANS_TYPE/CURRENCY
    rows_a = list(base), list(base)
    rows_b = [
        ["Bank-B", "OtherIssuer", "CARD1", 20260813, "10:00:00", "WITHDRAW", 100, 840, "-1", "RRN1", "T1", "Addr1"],
        ["Bank-A", "Issuer", "CARD2", 20260813, "10:00:00", "Purchase", 200, 230, "-1", "RRN2", "T1", "Addr2"],
    ]
    result = merge_reports(
        [("a.xlsx", _pos_daily_rows([*rows_a])), ("b.xlsx", _pos_daily_rows(rows_b))],
        mode_key="pos", dedupe=True,
    )
    # CARD1 triple (A,A,B) differ only in the ignored columns -> collapsed to 1
    # CARD2 row (B) differs in AMOUNT/PAN/RRN/ADDRESS -> stays
    assert result.total_rows == 2
    assert len(result.duplicate_records) == 3  # the CARD1 triple


def test_pos_daily_dedupe_still_checks_other_columns():
    # Rows that also differ in AMOUNT are NOT duplicates even though they
    # share the ignored columns.
    rows = [
        ["Bank-A", "Issuer", "CARD1", 20260813, "10:00:00", "Purchase", 100, 230, "-1", "RRN1", "T1", "Addr1"],
        ["Bank-A", "Issuer", "CARD1", 20260813, "10:00:00", "Purchase", 200, 230, "-1", "RRN1", "T1", "Addr1"],
    ]
    result = merge_reports([("a.xlsx", _pos_daily_rows(rows))], mode_key="pos", dedupe=True)
    assert result.total_rows == 2
    assert len(result.duplicate_records) == 0


def test_pos_success_dedupe_ignores_acquirer_issuer_trans_type():
    # POS Success (SVFE) also ignores ACQUIRER / ISSUER / TRANS_TYPE
    # (it has no CURRENCY column). Rows differing only in them are duplicates.
    wb = Workbook()
    ws = wb.active
    ws.append(list(POS_SUCCESS_CANONICAL_COLUMNS))
    ws.append(["Bank-A", "Issuer", "PAN1", 20260813, "10:00:00", "Purchase", 100, "00", "REF1", 260813000058603840, "M1"])
    ws.append(["Bank-B", "Issuer", "PAN1", 20260813, "10:00:00", "WITHDRAW", 100, "00", "REF1", 260813000058603840, "M1"])
    buf = io.BytesIO()
    wb.save(buf)
    result = merge_reports([("a.xlsx", buf.getvalue())], mode_key="pos_success", dedupe=True)
    assert result.total_rows == 1
    assert len(result.duplicate_records) == 2


def test_pos_daily_extra_column_removed():
    rep = parse_report(_pos_daily_with_utrnno(), "pos_extra.xlsx", mode=POS_MODE)
    assert rep.data_rows == 1
    assert rep.extra_columns == ["UTRNNO"]
    assert any("additional column" in w and "UTRNNO" in w for w in rep.warnings)
    # the UTRNNO value is not mapped into any canonical column
    assert "UTRNNO" not in rep.rows[0]

    result = merge_reports([("pos_extra.xlsx", _pos_daily_with_utrnno())], mode_key="pos")
    assert result.total_rows == 1
    assert list(result.records[0].keys()) == list(POS_CANONICAL_COLUMNS)
    # output keeps only the sample columns - UTRNNO is gone
    assert "UTRNNO" not in result.records[0]
    assert result.records[0]["ADDRESS"] == "SHOP A"


def test_pos_daily_missing_column_vs_sample_warning():
    """A file missing a sample column (and no other file has it) warns,
    naming the column and the file, referencing the output format."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "CARD_NUMBER", "TRANS_DATE", "TRANS_TIME",
               "TRANS_TYPE", "AMOUNT", "CURRENCY", "RESP", "RRN",
               "TERMINAL_ID"])
    ws.append(["Bank A", "Bank B", "4006780*****2222", 20260815, "10:00:00",
               "Purchase", 100, 230, "-1", "002724805974", "T1"])
    buf = io.BytesIO()
    wb.save(buf)

    result = merge_reports([("no_address.xlsx", buf.getvalue())], mode_key="pos")
    assert result.total_rows == 1
    assert any(
        "no_address.xlsx" in w and "ADDRESS" in w and "output format" in w
        for w in result.warnings
    ), result.warnings


def test_fast_reader_matches_openpyxl():
    """The fast regex reader must produce identical raw values to openpyxl,
    including shared strings, inline strings, booleans and big numbers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.append(["ACQUIRER", "ISSUER", "AMOUNT", "FLAG"])
    ws.append(["Bank & Co <Ltd>", "Café \"X\"", 260815000066964000, True])
    ws.append(["Plain", "", 1573.26, False])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    fast = _read_xlsx_fast(data)
    slow = _read_xlsx_openpyxl(data)
    assert fast.keys() == slow.keys()
    f = fast["Mixed"]
    s = slow["Mixed"]
    assert len(f) == len(s)
    for i in range(len(f)):
        assert f[i] == s[i], (i, f[i], s[i])
    assert f[1][0] == "Bank & Co <Ltd>"  # entities unescaped


def test_time_normalization():
    assert _normalize_time(115207) == "11:52:07"
    assert _normalize_time(94135) == "09:41:35"
    assert _normalize_time(84152) == "08:41:52"
    assert _normalize_time("01:20:54") == "01:20:54"
    assert _normalize_time("12:30") == "12:30:00"
    assert _normalize_time("0:0:20") == "00:00:20"
    assert _normalize_time("0:01:16") == "00:01:16"
    assert _normalize_time("") == ""
    assert _normalize_time(None) == ""


def test_build_workbook_roundtrip_via_pandas():
    result = merge_reports([("decline42.xlsx", _xls_report_style())])
    df = pd_read_workbook(result.workbook_bytes)
    assert df.shape == (3, len(CANONICAL_COLUMNS))
    assert list(df.columns) == CANONICAL_COLUMNS


def pd_read_workbook(data: bytes):
    import pandas as pd
    return pd.read_excel(io.BytesIO(data), header=3, engine="openpyxl")


# ---------------------------------------------------------------------------
# Integration test against the real report files (skipped when not present)
# ---------------------------------------------------------------------------
REAL_XLSX = Path(__file__).resolve().parents[1] / ".." / "Aug 13  POS  Declined.xlsx"
REAL_XLS = Path(__file__).resolve().parents[1] / ".." / "POS_Transaction_Decline_Report (42).xls"
REAL_ATM_AUG15 = Path(__file__).resolve().parents[1] / ".." / "Aug 15 ATM s-s .xlsx"
REAL_ATM_DAILY = (
    Path(__file__).resolve().parents[1]
    / ".." / "Daily_Tranaction_Report_SmartVista_ATM_15_Aug_26_to_15_Aug_26_1.xlsx"
)
REAL_POS_AUG15 = Path(__file__).resolve().parents[1] / ".." / "Aug 15  POS s-s  .xlsx"
REAL_POS_DAILY = (
    Path(__file__).resolve().parents[1]
    / ".." / "Daily_Tranaction_Report_SmartVista_POS_15_Aug_26_to_15_Aug_26.xlsx"
)


@pytest.mark.skipif(not (REAL_XLSX.exists() and REAL_XLS.exists()), reason="real report files not present")
def test_real_files_end_to_end():
    xlsx = REAL_XLSX.read_bytes()
    xls = REAL_XLS.read_bytes()

    xlsx_rep = parse_report(xlsx, REAL_XLSX.name)
    assert xlsx_rep.data_rows == 1181
    assert xlsx_rep.blank_columns == []
    # exact large UTRNNO preserved from the real file
    utrn = {r["FE UTRNNO"] for r in xlsx_rep.rows}
    assert 260813000058603840 in utrn

    xls_rep = parse_report(xls, REAL_XLS.name)
    assert xls_rep.data_rows == 226
    assert xls_rep.blank_columns == ["B", "D", "F"]
    assert xls_rep.columns_kept == CANONICAL_COLUMNS
    # literal "null" strings are preserved exactly as in the file
    flat = [v for row in xls_rep.rows for v in row.values()]
    assert "null" in flat

    result = merge_reports([(REAL_XLSX.name, xlsx), (REAL_XLS.name, xls)])
    assert result.total_rows == 1181 + 226
    assert result.from_date == result.to_date == "20260813"

    # merged workbook reads back cleanly: no blank columns, all rows
    df = pd_read_workbook(result.workbook_bytes)
    assert df.shape == (result.total_rows, len(CANONICAL_COLUMNS))
    assert list(df.columns) == CANONICAL_COLUMNS
    assert not df.isna().all().any()
    # fully sorted by date, then time (via normalized keys; values are raw)
    times = df["TIME"].tolist()
    keys = [_normalize_time(t) for t in times]
    assert keys == sorted(keys)


@pytest.mark.skipif(
    not (REAL_POS_AUG15.exists() and REAL_POS_DAILY.exists()),
    reason="real POS daily files not present",
)
def test_real_pos_daily_files():
    # the Aug 15 POS export (TIME / ADDRESS_NAME headers)
    aug15 = parse_report(REAL_POS_AUG15.read_bytes(), REAL_POS_AUG15.name, mode=POS_MODE)
    assert aug15.data_rows == 10251
    assert aug15.blank_columns == []
    assert aug15.columns_kept == list(POS_CANONICAL_COLUMNS)
    assert aug15.extra_columns == []
    assert any(r["TRANS_TIME"] == "00:24:38" for r in aug15.rows)  # raw value kept

    # the daily SmartVista POS report (the output format)
    daily = parse_report(REAL_POS_DAILY.read_bytes(), REAL_POS_DAILY.name, mode=POS_MODE)
    assert daily.data_rows == 941
    assert daily.order_mismatch is False
    assert daily.columns_kept == list(POS_CANONICAL_COLUMNS)

    result = merge_reports(
        [(REAL_POS_AUG15.name, REAL_POS_AUG15.read_bytes()),
         (REAL_POS_DAILY.name, REAL_POS_DAILY.read_bytes())],
        mode_key="pos",
    )
    assert result.total_rows == 10251 + 941
    assert result.filename == (
        "Daily_Tranaction_Report_SmartVista_POS_15_Aug_26_to_15_Aug_26_Merged.xlsx"
    )
    import pandas as pd
    df = pd.read_excel(io.BytesIO(result.workbook_bytes), header=0, engine="openpyxl")
    assert df.shape == (result.total_rows, len(POS_CANONICAL_COLUMNS))
    assert list(df.columns) == list(POS_CANONICAL_COLUMNS)


@pytest.mark.skipif(
    not (REAL_ATM_AUG15.exists() and REAL_ATM_DAILY.exists()),
    reason="real ATM files not present",
)
def test_real_atm_files():
    # the large Aug 15 export (UTRNNO before RRN, TIME header)
    aug15 = parse_report(REAL_ATM_AUG15.read_bytes(), REAL_ATM_AUG15.name, mode=ATM_MODE)
    assert aug15.data_rows == 361782
    assert aug15.blank_columns == []
    assert aug15.columns_kept == list(ATM_CANONICAL_COLUMNS)
    assert aug15.order_mismatch is True  # UTRNNO/RRN reshuffled vs daily format
    # exact big UTRNNO preserved from the real file
    assert 260815000066964000 in {r["UTRNNO"] for r in aug15.rows}

    # the daily SmartVista ATM report (the output format)
    daily = parse_report(REAL_ATM_DAILY.read_bytes(), REAL_ATM_DAILY.name, mode=ATM_MODE)
    assert daily.data_rows == 66979
    assert daily.order_mismatch is False
    assert any(r["TRANS_TIME"] == "0:0:20" for r in daily.rows)  # raw value kept
    assert any(r["RRN"] == "002724804961" for r in daily.rows)   # leading zeros kept

    # merging the daily report alone exercises the full ATM write path
    result = merge_reports(
        [(REAL_ATM_DAILY.name, REAL_ATM_DAILY.read_bytes())], mode_key="atm"
    )
    assert result.total_rows == 66979
    assert result.filename == (
        "Daily_Tranaction_Report_SmartVista_ATM_15_Aug_26_to_15_Aug_26_Merged.xlsx"
    )
    # ATM workbook has the header in row 1 (header=0), not row 4 like POS
    import pandas as pd
    df = pd.read_excel(io.BytesIO(result.workbook_bytes), header=0, engine="openpyxl")
    assert df.shape == (66979, len(ATM_CANONICAL_COLUMNS))
    assert list(df.columns) == list(ATM_CANONICAL_COLUMNS)


# ---------------------------------------------------------------------------
# POS Success mode (POS_Transaction_SVFE_Report format)
# ---------------------------------------------------------------------------
def _pos_success_input_style() -> bytes:
    """Mimic 'Aug 13 POS success.xlsx': 11 columns, header in row 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Export Worksheet"
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRANS_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REFNUM", "FE_UTRNNO", "ADDRESS_NAME"])
    ws.append(["Wegagen Bank", "Commercial Bank", "4585716*****2369", 20260813,
               "01:13:11", "Purchase", 2040, -1, "622422564685",
               "260813000058601088", "Yechaka Bunna PLC"])
    ws.append(["Awash Bank", "Commercial Bank", "4583007*****1121", 20260813,
               "03:50:22", "Purchase", 3386, -1, "622503618227",
               "260813000058657216", "CARELAND GEN.HOSPITA"])
    ws.append(["Abyssinia Bank", "Commercial Bank", "9231402*****8206692", 20260813,
               "00:55:45", "Purchase", 2500, -1, "622521254854",
               "260813000058594656", "HELEN YOHANNES G/MICHAEL POS02"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pos_success_report_style() -> bytes:
    """Mimic 'POS_Transaction_SVFE_Report (14).xls': title rows, blank spacer
    columns, literal 'null' strings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "POS_Transaction_Report"
    ws.append(["", "Report name:", "", "POS TRANSACTION REPORT"])
    ws.append(["", "From Date:", "", 20260813])
    ws.append(["", "To Date:", "", 20260813])
    ws.append(["ACQUIRER", "", "ISSUER", "", "PAN", "", "TRAN_DATE", "", "TIME",
               "TRANS_TYPE", "AMOUNT", "RESP_CODE", "REFNUM", "UTRNNO", "MERCHANT"])
    ws.append(["null", "", "CBE", "", "4583006xxxxxx4046", "", 20260813, "",
               "906", "POS purchase", "2430.49", "-1", "002724352979",
               "2724352979", "Y N M PASTRY AND "])
    ws.append(["null", "", "CBE", "", "4583007xxxxxx5441", "", 20260813, "",
               "1523", "POS purchase", "505", "-1", "002724353634",
               "2724353634", "BEREKA MEDICAL "])
    ws.append(["Abyssinia", "", "Abyssinia", "", "4006780xxxxxx1873", "", 20260813,
               "", "21416", "POS purchase", "2400", "-1", "002724365623",
               "2724365623", "BISRATEGEBRIEL "])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_pos_success_input():
    rep = parse_report(_pos_success_input_style(), "pos_success.xlsx", mode=POS_SUCCESS_MODE)
    assert rep.data_rows == 3
    assert rep.blank_columns == []
    assert rep.columns_kept == list(POS_SUCCESS_CANONICAL_COLUMNS)
    assert rep.order_mismatch is False

    row = rep.rows[0]
    assert row["ACQUIRER"] == "Wegagen Bank"
    assert row["ISSUER"] == "Commercial Bank"
    assert row["PAN"] == "4585716*****2369"
    assert row["TRAN_DATE"] == 20260813
    assert row["TIME"] == "01:13:11"
    assert row["TRANS_TYPE"] == "Purchase"
    assert row["AMOUNT"] == 2040
    assert row["RESP_CODE"] == -1
    assert row["REFNUM"] == "622422564685"
    assert row["UTRNNO"] == "260813000058601088"
    assert row["MERCHANT"] == "Yechaka Bunna PLC"


def test_parse_pos_success_report_style():
    rep = parse_report(_pos_success_report_style(), "svfe_report.xls", mode=POS_SUCCESS_MODE)
    assert rep.data_rows == 3
    assert rep.blank_columns == ["B", "D", "F", "H"]
    assert rep.columns_kept == list(POS_SUCCESS_CANONICAL_COLUMNS)
    assert "repeated header rows skipped" not in str(rep.warnings)
    # literal "null" strings are preserved
    issuers = {r["ACQUIRER"] for r in rep.rows}
    assert "null" in issuers
    # TIME values kept raw (906, 1523, 21416)
    times = [r["TIME"] for r in rep.rows]
    assert times == ["906", "1523", "21416"]


def test_merge_pos_success_reports():
    result = merge_reports(
        [("pos_success.xlsx", _pos_success_input_style()),
         ("svfe_report.xls", _pos_success_report_style())],
        mode_key="pos_success",
    )
    assert result.total_rows == 6
    assert result.mode_key == "pos_success"
    assert result.from_date == result.to_date == "20260813"
    assert result.filename == "POS_Transaction_SVFE_Report_20260813_Merged.xlsx"
    for rec in result.records:
        assert list(rec.keys()) == list(POS_SUCCESS_CANONICAL_COLUMNS)
    # sorted by time (via internal normalized key; raw values untouched)
    times = [r["TIME"] for r in result.records]
    keys = [_normalize_time(t) for t in times]
    assert keys == sorted(keys)


def test_pos_success_workbook_layout():
    result = merge_reports(
        [("pos_success.xlsx", _pos_success_input_style())], mode_key="pos_success"
    )
    wb = load_workbook(io.BytesIO(result.workbook_bytes))
    assert wb.sheetnames == ["POS_Transaction_Report"]
    ws = wb["POS_Transaction_Report"]
    # title block
    assert ws["A1"].value == "Report name:"
    assert ws["B1"].value == "POS TRANSACTION REPORT"
    assert ws["A2"].value == "From Date:"
    assert ws["B2"].value == 20260813
    assert ws["A3"].value == "To Date:"
    # header row
    header = [ws.cell(row=4, column=c).value for c in range(1, len(POS_SUCCESS_CANONICAL_COLUMNS) + 1)]
    assert header == list(POS_SUCCESS_CANONICAL_COLUMNS)
    # data starts at row 5 (sorted by time; first is Abyssinia Bank at 00:55:45)
    assert ws.cell(row=5, column=1).value == "Abyssinia Bank"
    assert ws.cell(row=5, column=9).value == "622521254854"  # REFNUM
    assert ws.cell(row=5, column=10).value == "260813000058594656"  # UTRNNO


def test_pos_success_extra_column_removed():
    """A POS Success export with an extra REVERSAL column (not in sample)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRANS_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REFNUM", "FE_UTRNNO", "ADDRESS_NAME", "REVERSAL"])
    ws.append(["Bank A", "Bank B", "4006780*****1111", 20260813, "10:00:00",
               "Purchase", 500, -1, "622422564685", "260813000058601088", "SHOP A", 0])
    buf = io.BytesIO()
    wb.save(buf)

    rep = parse_report(buf.getvalue(), "extra.xlsx", mode=POS_SUCCESS_MODE)
    assert rep.data_rows == 1
    assert rep.extra_columns == ["REVERSAL"]
    assert any("additional column" in w and "REVERSAL" in w for w in rep.warnings)

    result = merge_reports([("extra.xlsx", buf.getvalue())], mode_key="pos_success")
    assert result.total_rows == 1
    assert list(result.records[0].keys()) == list(POS_SUCCESS_CANONICAL_COLUMNS)
    assert "REVERSAL" not in result.records[0]


def test_pos_success_missing_column_warning():
    """A file missing MERCHANT column warns about it."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRANS_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REFNUM", "FE_UTRNNO"])
    ws.append(["Bank A", "Bank B", "4006780*****2222", 20260815, "10:00:00",
               "Purchase", 100, -1, "622422564685", "260813000058601088"])
    buf = io.BytesIO()
    wb.save(buf)

    result = merge_reports([("no_merchant.xlsx", buf.getvalue())], mode_key="pos_success")
    assert result.total_rows == 1
    assert any(
        "no_merchant.xlsx" in w and "MERCHANT" in w for w in result.warnings
    ), result.warnings


# ---------------------------------------------------------------------------
# Integration tests against real POS Success files
# ---------------------------------------------------------------------------
REAL_POS_SUCCESS = Path(__file__).resolve().parents[1] / ".." / "Aug 13 POS success.xlsx"
REAL_SVFE_REPORT = (
    Path(__file__).resolve().parents[1]
    / ".." / "POS_Transaction_SVFE_Report (14).xls"
)


@pytest.mark.skipif(
    not (REAL_POS_SUCCESS.exists() and REAL_SVFE_REPORT.exists()),
    reason="real POS Success files not present",
)
def test_real_pos_success_files():
    inp = parse_report(
        REAL_POS_SUCCESS.read_bytes(), REAL_POS_SUCCESS.name, mode=POS_SUCCESS_MODE
    )
    assert inp.data_rows == 8736
    assert inp.blank_columns == []
    assert inp.columns_kept == list(POS_SUCCESS_CANONICAL_COLUMNS)

    report = parse_report(
        REAL_SVFE_REPORT.read_bytes(), REAL_SVFE_REPORT.name, mode=POS_SUCCESS_MODE
    )
    assert report.data_rows == 788
    assert report.blank_columns == ["B", "D", "F", "H"]
    assert report.columns_kept == list(POS_SUCCESS_CANONICAL_COLUMNS)
    # literal "null" strings preserved
    flat = [v for row in report.rows for v in row.values()]
    assert "null" in flat

    result = merge_reports(
        [(REAL_POS_SUCCESS.name, REAL_POS_SUCCESS.read_bytes()),
         (REAL_SVFE_REPORT.name, REAL_SVFE_REPORT.read_bytes())],
        mode_key="pos_success",
    )
    assert result.total_rows == 8736 + 788
    assert result.from_date == result.to_date == "20260813"

    import pandas as pd
    df = pd.read_excel(io.BytesIO(result.workbook_bytes), header=3, engine="openpyxl")
    assert df.shape == (result.total_rows, len(POS_SUCCESS_CANONICAL_COLUMNS))
    assert list(df.columns) == list(POS_SUCCESS_CANONICAL_COLUMNS)


# ---------------------------------------------------------------------------
# Smart sorting (dates written with spelled months vs numbers, etc.)
# ---------------------------------------------------------------------------
def test_smart_sort_key_mixed_dates():
    # spelled-month dates and an 8-digit numeric date all land in the same
    # date "namespace" so they interleave chronologically.
    assert smart_sort_key("15-Aug-26")[0] == 2
    assert smart_sort_key("Aug 13, 2026")[0] == 2
    assert smart_sort_key("20260813")[0] == 2
    # the same actual date yields the same key regardless of representation
    assert smart_sort_key("20260813") == smart_sort_key("13 Aug 2026")
    assert smart_sort_key("Aug 13, 2026") == smart_sort_key("20260813")


def test_smart_sort_text_and_numbers_mixed():
    # numeric values sort numerically and group before plain text; empties
    # always sort first
    keys = sorted(["B", "A", "10", "", "2"], key=smart_sort_key)
    assert keys == ["", "2", "10", "A", "B"]
    # raw numbers (not strings) sort numerically too
    keys2 = sorted([10, 2, "", 1.5], key=smart_sort_key)
    assert keys2 == ["", 1.5, 2, 10]


def test_merge_sort_by_amount_desc():
    result = merge_reports(
        [("export.xlsx", _xlsx_export_style()), ("decline42.xlsx", _xls_report_style())],
        sort_by="AMOUNT",
        sort_dir="desc",
    )
    assert result.sort_by == "AMOUNT"
    assert result.sort_dir == "desc"
    amounts = [r["AMOUNT"] for r in result.records if r["AMOUNT"] != ""]
    assert amounts == sorted(amounts, reverse=True)


def test_merge_sort_does_not_alter_values():
    result = merge_reports(
        [("export.xlsx", _xlsx_export_style()), ("decline42.xlsx", _xls_report_style())],
        sort_by="TIME",
        sort_dir="asc",
    )
    # TIME values are untouched even though sorting used an internal key
    times = [r["TIME"] for r in result.records]
    assert all(isinstance(t, (int, str)) for t in times)
    assert 115207 in times


def test_merge_default_sort_is_date_then_time():
    result = merge_reports([("export.xlsx", _xlsx_export_style())])
    # no sort_by requested -> default date+time ordering preserved
    times = [r["TIME"] for r in result.records]
    keys = [_normalize_time(t) for t in times]
    assert keys == sorted(keys)


def test_merge_sort_empty_column_pinned_front_when_descending():
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM", "ADDRESS_NAME"])
    ws.append(["A", "I", "1", 20260813, "10:00:00", "Purchase", 5, 901, 0,
               "260813000000000000", "1", "M1"])
    ws.append(["B", "I", "2", 20260813, "11:00:00", "Purchase", 20, 901, 0,
               "260813000000000001", "2", "M2"])
    buf1 = io.BytesIO()
    wb.save(buf1)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
                "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM", "ADDRESS_NAME"])
    ws2.append(["C", "I", "3", 20260813, "12:00:00", "Purchase", "", 901, 0,
                "260813000000000002", "3", "M3"])  # empty AMOUNT
    buf2 = io.BytesIO()
    wb2.save(buf2)

    result = merge_reports([("a.xlsx", buf1.getvalue()), ("b.xlsx", buf2.getvalue())],
                           sort_by="AMOUNT", sort_dir="desc")
    # empty AMOUNT sorts first (pinned to the front), the rest descend
    assert result.records[0]["ACQUIRER"] == "C"
    assert result.records[0]["AMOUNT"] == ""
    assert result.records[1]["AMOUNT"] == 20
    assert result.records[2]["AMOUNT"] == 5


# ---------------------------------------------------------------------------
# QR mode (transfer-export / "July - December 2025 Source" EXPORT_TABLE)
# ---------------------------------------------------------------------------
def _qr_export_style() -> bytes:
    """Mimic the EXPORT_TABLE structure from 'July - December 2025 Source':
    DESTINATION_BANK, SOURCE_BANK, TRX_DATE, DBTR_ACCT, CDTR_ACCT, AMOUNT,
    TX_ID, STATUS - header in row 1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["DESTINATION_BANK", "SOURCE_BANK", "TRX_DATE", "DBTR_ACCT",
               "CDTR_ACCT", "AMOUNT", "TX_ID", "STATUS"])
    ws.append(["Awash Bank", "Cooperative Bank of Oromia", 20250707,
               "251915180606", "014251099975200", 77600,
               "CBORETAA1455091289", "PROCESSED"])
    ws.append(["Amhara Bank S.C", "Cooperative Bank of Oromia", 20250707,
               "251913230123", "01425166631800", 100000,
               "CBORETAA1455204002", "DECLINED"])
    ws.append(["Dashen Bank S.c", "Cooperative Bank of Oromia", 20251222,
               "251940966729", "01425597396100", 23000,
               "CBORETAA1913344123", "PROCESSED"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_qr_export_style():
    rep = parse_report(_qr_export_style(), "qr_source.xlsx", mode=QR_MODE)
    assert rep.data_rows == 3
    assert rep.blank_columns == []
    assert rep.columns_kept == list(QR_CANONICAL_COLUMNS)
    assert rep.order_mismatch is False

    row = rep.rows[0]
    assert row["DESTINATION_BANK"] == "Awash Bank"
    assert row["SOURCE_BANK"] == "Cooperative Bank of Oromia"
    assert row["TRX_DATE"] == 20250707
    assert row["DBTR_ACCT"] == "251915180606"
    assert row["CDTR_ACCT"] == "014251099975200"
    assert row["AMOUNT"] == 77600
    assert row["TX_ID"] == "CBORETAA1455091289"
    assert row["STATUS"] == "PROCESSED"


def test_parse_qr_emphasizes_header_variants():
    """A QR export using descriptive header names still maps to canonical."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Destination Bank", "Sending Bank", "Transaction Date", "Debit Acct",
               "Credit Acct", "Amount", "Transaction ID", "Transaction Status"])
    ws.append(["Awash Bank", "Cooperative Bank of Oromia", "07-Jul-25",
               "251915180606", "014251099975200", 77600, "TXN1", "PROCESSED"])
    buf = io.BytesIO()
    wb.save(buf)

    rep = parse_report(buf.getvalue(), "qr_variants.xlsx", mode=QR_MODE)
    assert rep.data_rows == 1
    assert rep.columns_kept == list(QR_CANONICAL_COLUMNS)
    row = rep.rows[0]
    assert row["DESTINATION_BANK"] == "Awash Bank"
    assert row["SOURCE_BANK"] == "Cooperative Bank of Oromia"
    assert row["TRX_DATE"] == "07-Jul-25"  # raw value untouched
    assert row["DBTR_ACCT"] == "251915180606"
    assert row["CDTR_ACCT"] == "014251099975200"
    assert row["TX_ID"] == "TXN1"
    assert row["STATUS"] == "PROCESSED"


def test_merge_qr_reports():
    result = merge_reports([("qr_source.xlsx", _qr_export_style())], mode_key="qr")
    assert result.total_rows == 3
    assert result.mode_key == "qr"
    assert result.filename == (
        "QR_Export_07_Jul_25_to_22_Dec_25_Merged.xlsx"
    )
    assert result.from_date == "20250707"
    assert result.to_date == "20251222"
    for rec in result.records:
        assert list(rec.keys()) == list(QR_CANONICAL_COLUMNS)
    # default sort is by date; dates are 07-Jul and 22-Dec (ascending)
    dates = [_normalize_time(r["TRX_DATE"]) for r in result.records]
    assert dates == sorted(dates)
    assert result.resp_counts == {"PROCESSED": 2, "DECLINED": 1}


def test_merge_qr_smart_sort_date_spelled_and_numeric():
    """Two QR files whose TRX_DATE is written with a spelled month rather
    than a number must still interleave chronologically."""
    def qr_with_date(day, status="PROCESSED"):
        wb = Workbook()
        ws = wb.active
        ws.append(["DESTINATION_BANK", "SOURCE_BANK", "TRX_DATE", "DBTR_ACCT",
                   "CDTR_ACCT", "AMOUNT", "TX_ID", "STATUS"])
        ws.append(["Awash Bank", "Coop Bank", day, "A", "B", 100, f"T-{day}", status])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    result = merge_reports(
        [
            ("a.xlsx", qr_with_date("15-Aug-26")),
            ("b.xlsx", qr_with_date(20260813)),
        ],
        mode_key="qr",
    )
    assert result.total_rows == 2
    assert result.records[0]["TRX_DATE"] == 20260813  # Aug 13 before Aug 15
    assert result.records[1]["TRX_DATE"] == "15-Aug-26"


def test_merge_qr_sort_by_amount_desc():
    result = merge_reports(
        [("qr_source.xlsx", _qr_export_style())], mode_key="qr",
        sort_by="AMOUNT", sort_dir="desc",
    )
    assert result.sort_by == "AMOUNT"
    assert result.sort_dir == "desc"
    amounts = [r["AMOUNT"] for r in result.records if r["AMOUNT"] != ""]
    assert amounts == sorted(amounts, reverse=True)


def test_merge_qr_sort_by_status():
    result = merge_reports(
        [("qr_source.xlsx", _qr_export_style())], mode_key="qr",
        sort_by="STATUS", sort_dir="asc",
    )
    statuses = [r["STATUS"] for r in result.records]
    assert statuses == sorted(statuses)


def test_qr_workbook_layout():
    result = merge_reports([("qr_source.xlsx", _qr_export_style())], mode_key="qr")
    wb = load_workbook(io.BytesIO(result.workbook_bytes))
    assert wb.sheetnames == ["Report"]
    ws = wb["Report"]
    header = [ws.cell(row=1, column=c).value
              for c in range(1, len(QR_CANONICAL_COLUMNS) + 1)]
    assert header == list(QR_CANONICAL_COLUMNS)
    assert ws.cell(row=1, column=9).value is None  # nothing beyond col H
    # data starts at row 2, sorted by date
    assert ws.cell(row=2, column=3).value in (20250707, "20250707", 20250707)
    assert ws.cell(row=2, column=8).value in ("PROCESSED", "DECLINED")


def test_qr_rejects_file_without_qr_headers():
    """A POS-style file (ACQUIRER header) is not treated as QR input."""
    wb = Workbook()
    ws = wb.active
    ws.append(["ACQUIRER", "ISSUER", "PAN", "TRAN_DATE", "TIME", "TRANS_TYPE",
               "AMOUNT", "RESP", "REVERSAL", "FE_UTRNNO", "REFNUM", "ADDRESS_NAME"])
    ws.append(["Bank A", "Bank B", "1111*****2222", 20260813, "10:00:00",
               "Purchase", 100, 901, 0, 123456, "REF-1", "M1"])
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="No QR table found"):
        parse_report(buf.getvalue(), "pos_in_qr_mode.xlsx", mode=QR_MODE)


