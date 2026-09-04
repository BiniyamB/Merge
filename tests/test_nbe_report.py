"""Unit tests for nbe_report module."""

import io
import openpyxl
import pandas as pd
import pytest

from nbe_report import (
    STANDARD_NBE_BANKS,
    build_nbe_report_excel,
    generate_nbe_report,
    normalize_nbe_bank,
)


def test_normalize_nbe_bank():
    """Verify alias mapping for issuer & acquirer bank names."""
    assert normalize_nbe_bank("Abay") == "Abay Bank"
    assert normalize_nbe_bank("Abay Bank") == "Abay Bank"
    assert normalize_nbe_bank("Abyssinia") == "BOA"
    assert normalize_nbe_bank("Abyssinia Bank") == "BOA"
    assert normalize_nbe_bank("BOA") == "BOA"
    assert normalize_nbe_bank("Commercial Bank") == "CBE"
    assert normalize_nbe_bank("Commercial Bank of Ethiopia") == "CBE"
    assert normalize_nbe_bank("CBE") == "CBE"
    assert normalize_nbe_bank("Hibret Bank") == "United Bank"
    assert normalize_nbe_bank("UB") == "United Bank"
    assert normalize_nbe_bank("Global") == "Global Bank"
    assert normalize_nbe_bank("Debub Bank") == "Global Bank"
    assert normalize_nbe_bank("Berhan Bank") == "Birhan Bank"
    assert normalize_nbe_bank("Tsedey Bank") == "Tseday Bank"
    assert normalize_nbe_bank(None) == ""


def test_generate_nbe_report_pos():
    """Verify POS NBE Report filtering and aggregation."""
    records = [
        # Match (POS purchase, RESP -1)
        {"TRANS_TYPE": "POS purchase", "RESP": "-1", "ISSUER": "Commercial Bank", "ACQUIRER": "Abyssinia Bank", "AMOUNT": 1000.0},
        {"TRANS_TYPE": "purchase", "RESP": -1.0, "ISSUER": "Awash Bank", "ACQUIRER": "CBE", "AMOUNT": 500.0},
        # Non-matching response code
        {"TRANS_TYPE": "POS purchase", "RESP": "901", "ISSUER": "Awash Bank", "ACQUIRER": "CBE", "AMOUNT": 200.0},
        # Non-matching trans type
        {"TRANS_TYPE": "POS balance inquiry", "RESP": "-1", "ISSUER": "Awash Bank", "ACQUIRER": "CBE", "AMOUNT": 0.0},
    ]

    df = generate_nbe_report(records, mode_key="pos")
    assert not df.empty
    assert "BANKS" in df.columns
    assert "PURCHASE As Issuer (Count)" in df.columns
    assert "PURCHASE As Acquirer (Amount ETB)" in df.columns

    # CBE as Issuer: 1 txn (1000.0)
    cbe_row = df[df["BANKS"] == "CBE"].iloc[0]
    assert cbe_row["PURCHASE As Issuer (Count)"] == 1
    assert cbe_row["PURCHASE As Issuer (Amount ETB)"] == 1000.0
    assert cbe_row["PURCHASE As Acquirer (Count)"] == 1
    assert cbe_row["PURCHASE As Acquirer (Amount ETB)"] == 500.0

    # Total row
    tot_row = df[df["BANKS"] == "Total"].iloc[0]
    assert tot_row["PURCHASE As Issuer (Count)"] == 2
    assert tot_row["PURCHASE As Issuer (Amount ETB)"] == 1500.0
    assert tot_row["PURCHASE As Acquirer (Count)"] == 2
    assert tot_row["PURCHASE As Acquirer (Amount ETB)"] == 1500.0


def test_generate_nbe_report_atm():
    """Verify ATM NBE Report filtering and aggregation."""
    records = [
        {"TRANS_TYPE": "ATM Cash withdrawal", "RESP": "-1.0", "ISSUER": "Abay Bank", "ACQUIRER": "Wegagen Bank", "AMOUNT": 400.0},
        {"TRANS_TYPE": "cash withdrawal", "RESP": -1, "ISSUER": "Wegagen Bank", "ACQUIRER": "BOA", "AMOUNT": 600.0},
    ]

    df = generate_nbe_report(records, mode_key="atm")
    tot_row = df[df["BANKS"] == "Total"].iloc[0]
    assert tot_row["CASH WITHDRAWAL As Issuer (Count)"] == 2
    assert tot_row["CASH WITHDRAWAL As Issuer (Amount ETB)"] == 1000.0
    assert tot_row["CASH WITHDRAWAL As Acquirer (Count)"] == 2
    assert tot_row["CASH WITHDRAWAL As Acquirer (Amount ETB)"] == 1000.0


def test_build_nbe_report_excel():
    """Verify Excel workbook generation."""
    records = [
        {"TRANS_TYPE": "POS purchase", "RESP": "-1", "ISSUER": "CBE", "ACQUIRER": "BOA", "AMOUNT": 1250.50},
    ]
    df = generate_nbe_report(records, mode_key="pos")
    excel_bytes = build_nbe_report_excel(df, mode_key="pos")
    assert len(excel_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    assert "NBE REPORT" in ws["A1"].value
    assert ws["B2"].value == "BANKS"
    assert ws["C2"].value == "PURCHASE As Issuer"
