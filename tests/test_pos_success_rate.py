"""Unit tests for pos_success_rate module."""

import io
import openpyxl
import pandas as pd
import pytest

from pos_success_rate import (
    CARDHOLDER_RC,
    build_pos_success_rate_excel,
    generate_pos_success_rate_report,
    parse_rc,
)


def test_parse_rc():
    """Verify response code string normalization."""
    assert parse_rc("-1") == "-1"
    assert parse_rc("-1.0") == "-1"
    assert parse_rc(0) == "-1"
    assert parse_rc("901") == "901"
    assert parse_rc("901.0") == "901"
    assert parse_rc(915.0) == "915"
    assert parse_rc(None) == ""


def test_generate_pos_success_rate_report():
    """Verify Issuer-based POS Success Rate report generation and formulas."""
    records = [
        # Issuer 1: Abyssinia -> BOA
        {"ISSUER": "Abyssinia Bank", "RESP": "-1"},  # Success
        {"ISSUER": "Abyssinia Bank", "RESP": "-1"},  # Success
        {"ISSUER": "Abyssinia Bank", "RESP": "901"},  # Cardholder decline (901)
        {"ISSUER": "Abyssinia Bank", "RESP": "915"},  # Cardholder decline (915)
        {"ISSUER": "Abyssinia Bank", "RESP": "801"},  # System decline (801)
        # Issuer 2: Commercial Bank -> CBE
        {"ISSUER": "Commercial Bank", "RESP": "-1"},  # Success
        {"ISSUER": "Commercial Bank", "RESP": "802"},  # System decline (802)
    ]

    matrix_df, desc_df = generate_pos_success_rate_report(records)

    assert not matrix_df.empty
    assert "BOA" in matrix_df.columns
    assert "CBE" in matrix_df.columns
    assert "Total" in matrix_df.columns

    boa_col = matrix_df.set_index("RC/BANK NAME")["BOA"]

    # BOA:
    # 801: 1, 901: 1, 915: 1
    # Total Decline = 3 (801 + 901 + 915)
    # Successful Pos T = 2 (-1, -1)
    # card holder rel dec = 2 (901 + 915)
    # total succ = 4 (2 + 2)
    # total pos t = 5 (3 + 2)
    # success rate = 4/5 = 0.8 (80%)

    assert int(boa_col["Total Decline"]) == 3
    assert int(boa_col["Successful Pos T"]) == 2
    assert int(boa_col["card holder rel dec"]) == 2
    assert int(boa_col["total succ"]) == 4
    assert int(boa_col["total pos t"]) == 5
    assert float(boa_col["success rate"]) == 0.8

    # CBE:
    # 802: 1
    # Total Decline = 1
    # Successful Pos T = 1
    # card holder rel dec = 0
    # total succ = 1
    # total pos t = 2
    # success rate = 1/2 = 0.5 (50%)
    cbe_col = matrix_df.set_index("RC/BANK NAME")["CBE"]
    assert int(cbe_col["Total Decline"]) == 1
    assert int(cbe_col["Successful Pos T"]) == 1
    assert int(cbe_col["card holder rel dec"]) == 0
    assert int(cbe_col["total succ"]) == 1
    assert int(cbe_col["total pos t"]) == 2
    assert float(cbe_col["success rate"]) == 0.5

    # Total Column
    tot_col = matrix_df.set_index("RC/BANK NAME")["Total"]
    assert int(tot_col["Total Decline"]) == 4
    assert int(tot_col["Successful Pos T"]) == 3
    assert int(tot_col["card holder rel dec"]) == 2
    assert int(tot_col["total succ"]) == 5
    assert int(tot_col["total pos t"]) == 7
    assert float(tot_col["success rate"]) == round(5 / 7, 4)

    # Response code lookup table
    assert not desc_df.empty
    assert "Response Code" in desc_df.columns
    assert "Description" in desc_df.columns
    assert "Remark" in desc_df.columns


def test_build_pos_success_rate_excel():
    """Verify POS Success Rate Excel output including formulas and tier color styling."""
    records = [
        {"ISSUER": "Awash Bank", "RESP": "-1"},
        {"ISSUER": "Awash Bank", "RESP": "901"},
    ]
    matrix_df, desc_df = generate_pos_success_rate_report(records)
    excel_bytes = build_pos_success_rate_excel(matrix_df, desc_df)

    assert len(excel_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    assert "Pos Transaction Decline Response & Success Rate Summary" in ws["A1"].value
    assert ws["A3"].value == "RC/BANK NAME"

    # Find row indices by row label in column A
    row_map = {}
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            row_map[str(val).strip()] = r

    assert "Total Decline" in row_map
    assert "card holder rel dec" in row_map
    assert "total succ" in row_map
    assert "total pos t" in row_map
    assert "success rate" in row_map

    # Check AWASH column (Col B / Col 2) formulas
    tot_dec_row = row_map["Total Decline"]
    tot_succ_row = row_map["total succ"]
    tot_pos_row = row_map["total pos t"]
    succ_rate_row = row_map["success rate"]

    assert str(ws.cell(row=tot_dec_row, column=2).value).startswith("=SUM(")
    assert str(ws.cell(row=tot_succ_row, column=2).value).startswith("=")
    assert str(ws.cell(row=tot_pos_row, column=2).value).startswith("=")
    assert str(ws.cell(row=succ_rate_row, column=2).value).startswith("=IF(")

    # Check Total Column (Col C / Col 3) formulas
    assert str(ws.cell(row=tot_dec_row, column=3).value).startswith("=SUM(")
    assert str(ws.cell(row=succ_rate_row, column=3).value).startswith("=IF(")

    # Check success rate row formatting & fill color (Awash rate = 2/2 = 100% -> Green fill C6EFCE)
    rate_cell = ws.cell(row=succ_rate_row, column=2)
    assert rate_cell.number_format == "0.00%"
    assert rate_cell.fill.start_color.rgb == "00C6EFCE" or rate_cell.fill.start_color.rgb == "C6EFCE"

